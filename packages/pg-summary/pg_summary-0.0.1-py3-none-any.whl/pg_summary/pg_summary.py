#!/usr/bin/env python

from __future__ import annotations

import argparse
import getpass
import logging
import sys
import traceback
from pathlib import Path

import openpyxl
import psycopg2
from openpyxl.styles import Border, Color, Font, PatternFill, Side
from openpyxl.utils.cell import get_column_letter
from psycopg2.extras import DictCursor
from psycopg2.sql import SQL, Composable, Identifier, Literal

__title__ = "pg_summary"
__author__ = "Caleb Grant"
__url__ = "https://github.com/geocoug/pg-summary"
__author_email__ = "grantcaleb22@gmail.com"
__license__ = "GNU GPLv3"
__version__ = "0.0.1"
__description__ = "Create a summary of unique values for each column in a Postgres table or view and summarize results in an Excel workbook."  # noqa: E501

logging.basicConfig(
    level=logging.INFO,
    format="%(message)s",
    handlers=[logging.NullHandler()],
)
logger = logging.getLogger(__name__)

FONT = Font(bold=True, size=12)
FILL = PatternFill(patternType="solid", fgColor=Color("97b4c9"))
BORDER = Border(
    left=Side(border_style="thin", color=Color("000000")),
    right=Side(border_style="thin", color=Color("000000")),
    top=Side(border_style="thin", color=Color("000000")),
    bottom=Side(border_style="thin", color=Color("000000")),
)


class PostgresDB:
    """Base database object."""

    def __init__(
        self: PostgresDB,
        host: str,
        database: str,
        user: str,
        port: int = 5432,
        passwd: None | str = None,
    ) -> None:
        self.host = host
        self.port = port
        self.database = database
        self.user = user
        if passwd is not None:
            self.passwd = passwd
        else:
            self.passwd = self.get_password()
        self.in_transaction = False
        self.encoding = "UTF8"
        self.conn = None
        if not self.valid_connection():
            raise psycopg2.Error(f"Error connecting to {self!s}")

    def __repr__(self: PostgresDB) -> str:
        return f"{self.__class__.__name__}(host={self.host}, port={self.port}, database={self.database}, user={self.user})"  # noqa: E501

    def get_password(self):
        try:
            return getpass.getpass(
                f"The script {Path(__file__).name} wants the password for {self!s}: ",
            )
        except (KeyboardInterrupt, EOFError) as err:
            raise err

    def valid_connection(self: PostgresDB) -> bool:
        """Test the database connection."""
        logger.debug(f"Testing connection to {self!s}")
        try:
            self.open_db()
            logger.debug(f"Connected to {self!s}")
            return True
        except psycopg2.Error:
            return False
        finally:
            self.close()

    def open_db(self: PostgresDB) -> None:
        """Open a database connection."""

        def db_conn(db):
            """Return a database connection object."""
            return psycopg2.connect(
                host=str(db.host),
                database=str(db.database),
                port=db.port,
                user=str(db.user),
                password=str(db.passwd),
            )

        if self.conn is None:
            self.conn = db_conn(self)
            self.conn.set_session(autocommit=False)
        self.encoding = self.conn.encoding

    def cursor(self: PostgresDB):
        """Return the connection cursor."""
        self.open_db()
        return self.conn.cursor(cursor_factory=DictCursor)

    def close(self: PostgresDB) -> None:
        """Close the database connection."""
        self.rollback()
        if self.conn is not None:
            self.conn.close()
            self.conn = None

    def commit(self: PostgresDB) -> None:
        """Commit the current transaction."""
        if self.conn:
            self.conn.commit()
        self.in_transaction = False

    def rollback(self: PostgresDB) -> None:
        """Roll back the current transaction."""
        if self.conn is not None:
            self.conn.rollback()
        self.in_transaction = False

    def execute(self: PostgresDB, sql: str | Composable, params=None):
        """A shortcut to self.cursor().execute() that handles encoding.

        Handles insert, updates, deletes
        """
        self.in_transaction = True
        try:
            curs = self.cursor()
            if isinstance(sql, Composable):
                logger.debug(f"\n{sql.as_string(curs)}")
                curs.execute(sql)
            else:
                if params is None:
                    logger.debug(f"\n{sql}")
                    curs.execute(sql.encode(self.encoding))
                else:
                    logger.debug(f"\nSQL:\n{sql}\nParameters:\n{params}")
                    curs.execute(sql.encode(self.encoding), params)
        except Exception:
            self.rollback()
            raise
        return curs

    def rowdict(self: PostgresDB, sql: str | Composable, params=None) -> tuple:
        """Convert a cursor object to an iterable that.

        yields dictionaries of row data.
        """
        curs = self.execute(sql, params)
        headers = [d[0] for d in curs.description]

        def dict_row():
            """Convert a data row to a dictionary."""
            row = curs.fetchone()
            if row:
                if self.encoding:
                    r = [
                        (
                            c.decode(self.encoding, "backslashreplace")
                            if isinstance(c, bytes)
                            else c
                        )
                        for c in row
                    ]
                else:
                    r = row
                return dict(zip(headers, r, strict=True))
            return None

        return (iter(dict_row, None), headers, curs.rowcount)


class PgSummary:

    def __init__(
        self: PgSummary,
        host: str,
        database: str,
        user: str,
        table_or_view: str,
        passwd: None | str = None,
        port: int = 5432,
        schema: str = "public",
        outfile: Path = None,
        include_src: bool = False,
    ) -> None:
        """Create a summary of unique values for each column in a Postgres table or view and summarize results in an Excel workbook.

        Args:
            host (str): Name of the Postgres host.
            database (str): Name of the Postgres database.
            user (str): Name of the Postgres user.
            table_or_view (str): Name of the Postgres table or view to summarize.
            passwd (str, optional): Password for the Postgres user. If not defined, the user will be prompted for the password. Defaults to None.
            port (int, optional): Port number of the Postgres host. Defaults to 5432.
            schema (str, optional): Name of the Postgres schema containing the table or view to summarize. Defaults to "public".
            outfile (Path, optional): Path to the output Excel file containing summarized results. If not defined, the output file will be named `PgSummary_{table_or_view}.xlsx`. Defaults to None.
            include_src (bool, optional): Include the source table or view in the output Excel file as a seperate sheet. Defaults to False.

        Raises:
            ValueError: If the schema does not exist.
            ValueError: If the table or view does not exist.
            ValueError: If the table or view has no rows.

        Examples:
            >>> from pg_summary import PgSummary
            ...
            ... PgSummary(
            ...     host="localhost",
            ...     database="mydb",
            ...     user="username",
            ...     table_or_view="mytable",
            ...     schema="public",
            ...     include_src=True,
            ... ).summary()
        """  # noqa: E501
        self.db = PostgresDB(host, database, user, port, passwd)
        self.schema = schema
        self.table_or_view = table_or_view
        self.outfile = outfile
        self.include_src = include_src
        logger.debug(f"{self!s}")
        self._validate_args()

    def __repr__(self: PgSummary) -> str:
        return f"{self.__class__.__name__}(db={self.db!s}, schema={self.schema}, table_or_view={self.table_or_view}, outfile={self.outfile})"  # noqa: E501

    def _validate_args(self: PgSummary) -> None:
        """Validate the arguments."""
        if not self.table_or_view:
            raise ValueError("No table or view specified.")
        if not self.outfile:
            self.outfile = Path(f"{self.__class__.__name__}_{self.table_or_view}.xlsx")
            logger.debug(f"No output file specified. Using table name: {self.outfile}")
        if not self.outfile.suffix == ".xlsx":
            raise ValueError(
                f"Output file must be an Excel file. Received: {self.outfile}"
            )
        if not self._schema_exists():
            raise ValueError(f"Schema {self.schema} does not exist.")
        if not self._table_or_view_exists():
            raise ValueError(
                f"Table or view {self.schema}.{self.table_or_view} does not exist."
            )
        if not self._table_or_view_has_rows():
            raise ValueError(
                f"Table or view {self.schema}.{self.table_or_view} has no rows."
            )

    def _schema_exists(self: PgSummary) -> bool:
        """Validate the schema exists."""
        sql = SQL(
            """
            SELECT schema_name
            FROM information_schema.schemata
            WHERE schema_name = {schema};
        """
        ).format(schema=Literal(self.schema))
        return self.db.execute(sql).rowcount > 0

    def _table_or_view_exists(self: PgSummary) -> bool:
        """Validate the table or view exists."""
        sql = SQL(
            """
            SELECT * FROM (
                SELECT table_schema, table_name
                FROM information_schema.tables
                UNION
                SELECT table_schema, table_name
                FROM information_schema.views
            ) t WHERE table_schema = {schema} AND table_name = {table};
        """
        ).format(schema=Literal(self.schema), table=Literal(self.table_or_view))
        return self.db.execute(sql).rowcount > 0

    def _table_column_exists(self: PgSummary, column: str) -> bool:
        """Validate the column exists in the table or view."""
        sql = SQL(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = {schema}
            AND table_name = {table}
            AND column_name = {column};
        """
        ).format(
            schema=Literal(self.schema),
            table=Literal(self.table_or_view),
            column=Literal(column),
        )
        return self.db.execute(sql).rowcount > 0

    def _table_or_view_has_rows(self: PgSummary) -> bool:
        """Validate the table or view has data rows."""
        sql = SQL(
            """
            SELECT * FROM {schema}.{table} LIMIT 1;
        """
        ).format(schema=Identifier(self.schema), table=Identifier(self.table_or_view))
        return self.db.execute(sql).rowcount > 0

    def _column_has_rows(self: PgSummary, column: str) -> bool:
        """Validate the column has data rows."""
        sql = SQL(
            """
            SELECT {column} FROM {schema}.{table} WHERE {column} IS NOT NULL LIMIT 1;
        """
        ).format(
            column=Identifier(column),
            schema=Identifier(self.schema),
            table=Identifier(self.table_or_view),
        )
        return self.db.execute(sql).rowcount > 0

    def get_column_names(self: PgSummary) -> list:
        """Return a list of column names in a table or view."""
        sql = SQL(
            """
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = {schema}
            AND table_name = {table}
            ORDER BY ordinal_position;
        """
        ).format(schema=Literal(self.schema), table=Literal(self.table_or_view))
        return [c["column_name"] for c in self.db.rowdict(sql)[0]]

    def get_column_dtype(self: PgSummary, column: str) -> str:
        """Return the data type of a column."""
        if not self._table_column_exists(column):
            raise ValueError(
                f"Column {column} does not exist in {self.schema}.{self.table_or_view}."
            )
        sql = SQL(
            """
            SELECT 
                column_name,
                case
                    when character_maximum_length is null
                    then data_type
                    else data_type || '(' || character_maximum_length || ')'
                    end as data_type
            FROM information_schema.columns
            WHERE table_schema = {schema}
            AND table_name = {table}
            AND column_name = {column};
        """
        ).format(
            schema=Literal(self.schema),
            table=Literal(self.table_or_view),
            column=Literal(column),
        )
        data = self.db.rowdict(sql)
        if data[-1] == 0:
            raise ValueError(
                f"Column {column} does not exist in {self.schema}.{self.table_or_view}."
            )
        return next(data[0])["data_type"]

    def get_unique_column_values(self: PgSummary, column: str) -> tuple:
        """Return a list of unique values for a column.

        Args:
            column (str): Name of the column to query for unique values.

        Returns:
            tuple: A tuple containing a list of unique values and the number of unique values.
        """
        if not self._table_column_exists(column):
            raise ValueError(
                f"Column {column} does not exist in {self.schema}.{self.table_or_view}."
            )
        sql = SQL(
            """
            SELECT DISTINCT {column}
            FROM {schema}.{table}
            WHERE {column} IS NOT NULL
            ORDER BY {column};
        """
        ).format(
            column=Identifier(column),
            schema=Identifier(self.schema),
            table=Identifier(self.table_or_view),
        )
        return self.db.rowdict(sql)

    def get_null_column_value_count(self: PgSummary, column: str) -> int:
        """Return the number of null values for a column.

        Args:
            column (str): Name of the column to query for null values.

        Returns:
            int: The number of null values for the column.
        """
        if not self._table_column_exists(column):
            raise ValueError(
                f"Column {column} does not exist in {self.schema}.{self.table_or_view}."
            )
        sql = SQL(
            """
            SELECT COUNT(*)
            FROM {schema}.{table}
            WHERE {column} IS NULL;
        """
        ).format(
            column=Identifier(column),
            schema=Identifier(self.schema),
            table=Identifier(self.table_or_view),
        )
        return next(self.db.rowdict(sql)[0])["count"]

    def get_table_row_count(self: PgSummary) -> int:
        """Return the number of rows in a table.

        Returns:
            int: The number of rows in the table.
        """
        sql = SQL(
            """
            SELECT COUNT(*)
            FROM {schema}.{table};
        """
        ).format(
            schema=Identifier(self.schema),
            table=Identifier(self.table_or_view),
        )
        return next(self.db.rowdict(sql)[0])["count"]

    def summary(self: PgSummary) -> None:
        """Create a summary of unique values for each column in a Postgres table or view and summarize results in an Excel workbook."""  # noqa: E501
        logger.info(f"Creating summary table for {self.schema}.{self.table_or_view}")
        wb = openpyxl.Workbook()
        sheet = wb.active
        if not sheet:
            sheet = wb.create_sheet()
        sheet.title = f"{self.__class__.__name__}"

        # Summary header
        sheet.cell(row=1, column=2).value = "Host"
        sheet.cell(row=1, column=2).font = FONT
        sheet.cell(row=2, column=2).value = self.db.host

        sheet.cell(row=1, column=3).value = "Database"
        sheet.cell(row=1, column=3).font = FONT
        sheet.cell(row=2, column=3).value = self.db.database

        sheet.cell(row=1, column=4).value = "Schema"
        sheet.cell(row=1, column=4).font = FONT
        sheet.cell(row=2, column=4).value = self.schema

        sheet.cell(row=1, column=5).value = "Table/View"
        sheet.cell(row=1, column=5).font = FONT
        sheet.cell(row=2, column=5).value = self.table_or_view

        sheet.cell(row=1, column=6).value = "Total Rows"
        sheet.cell(row=1, column=6).font = FONT
        sheet.cell(row=2, column=6).value = self.get_table_row_count()

        # Summary results
        for i, col in enumerate(self.get_column_names()):
            if i == 0:
                sheet.cell(row=4, column=i + 1).value = "# of unique values"
                sheet.cell(row=4, column=i + 1).font = FONT
                sheet.cell(row=5, column=i + 1).value = "# of null values"
                sheet.cell(row=5, column=i + 1).font = FONT
                sheet.cell(row=6, column=i + 1).value = "data type"
                sheet.cell(row=6, column=i + 1).font = FONT
                sheet.cell(row=7, column=i + 1).value = "column name"
                sheet.cell(row=7, column=i + 1).font = FONT
            # Number of unique values
            sheet.cell(row=4, column=i + 2).value = self.get_unique_column_values(col)[
                -1
            ]
            # Number of null values
            sheet.cell(row=5, column=i + 2).value = self.get_null_column_value_count(
                col
            )
            # Column data type
            sheet.cell(row=6, column=i + 2).value = self.get_column_dtype(col)
            # Column name
            sheet.cell(row=7, column=i + 2).value = col
            # Styling
            sheet.cell(row=7, column=i + 2).font = FONT
            sheet.cell(row=7, column=i + 2).border = BORDER
            sheet.cell(row=7, column=i + 2).fill = FILL
            # Unique column values
            for j, val in enumerate(self.get_unique_column_values(col)[0]):
                sheet.cell(row=j + 8, column=i + 2).value = val[col]
            # Turn on filtering for the column row
            sheet.auto_filter.ref = f"B7:{get_column_letter(i+2)}{j+8}"
            # Column A cell width
            # sheet.column_dimensions["A"].width = 20
        # Auto size all columns
        for col in sheet.columns:
            length = max(len(str(cell.value)) for cell in col)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = length
        # Write the source table or view to a new sheet
        if self.include_src:
            logger.debug(
                f"Writing source table or view to sheet {self.schema}.{self.table_or_view}"
            )
            sheet = wb.create_sheet(title=f"{self.schema}.{self.table_or_view}")
            sql = SQL("SELECT * FROM {schema}.{table};").format(
                schema=Identifier(self.schema),
                table=Identifier(self.table_or_view),
            )
            data, headers, rowcount = self.db.rowdict(sql)
            for i, header in enumerate(headers):
                sheet.cell(row=1, column=i + 1).value = header
                sheet.cell(row=1, column=i + 1).font = FONT
                sheet.cell(row=1, column=i + 1).border = BORDER
                sheet.cell(row=1, column=i + 1).fill = FILL
            for i, row in enumerate(data):
                for j, val in enumerate(row.values()):
                    sheet.cell(row=i + 2, column=j + 1).value = val
            # Auto size all columns
            for col in sheet.columns:
                length = max(len(str(cell.value)) for cell in col)
                sheet.column_dimensions[get_column_letter(col[0].column)].width = length
            # Turn on filtering for the column row
            sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        wb.save(self.outfile)
        logger.info(f"Summary complete. Saving output to {self.outfile}")


def clparser() -> argparse.Namespace:
    """Command line argument parser for Boilerplate."""
    parser = argparse.ArgumentParser(
        description=__description__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--version",
        action="version",
        version=f"%(prog)s {__version__}",
    )
    parser.add_argument(
        "--debug",
        dest="debug",
        action="store_true",
        help="enable debug mode.",
    )
    parser.add_argument(
        "-q",
        "--quiet",
        dest="quiet",
        action="store_true",
        help="suppress all console output.",
    )
    parser.add_argument(
        "-v",
        "--host",
        type=str,
        dest="host",
        required=True,
        help="Postgres host.",
    )
    parser.add_argument(
        "-p",
        "--port",
        type=int,
        default=5432,
        dest="port",
        required=False,
        help="Postgres port.",
    )
    parser.add_argument(
        "-u",
        "--user",
        type=str,
        dest="user",
        required=True,
        help="Postgres user.",
    )
    parser.add_argument(
        "-d",
        "--database",
        type=str,
        dest="database",
        required=True,
        help="Postgres database.",
    )
    parser.add_argument(
        "-s",
        "--schema",
        type=str,
        default="public",
        dest="schema",
        required=False,
        help="Postgres schema.",
    )
    parser.add_argument(
        "-t",
        "--table",
        type=str,
        dest="table",
        required=True,
        help="Postgres table or view.",
    )
    parser.add_argument(
        "-o",
        "--outfile",
        type=Path,
        dest="outfile",
        required=False,
        help="Output Excel file.",
    )
    parser.add_argument(
        "-i",
        "--include-src",
        dest="include_src",
        action="store_true",
        help="Include the source table or view in the output Excel file.",
    )
    return parser.parse_args()


def cli() -> None:
    """Command line interface for PgSummary."""
    args = clparser()
    if not args.quiet:
        logger.addHandler(logging.StreamHandler())
    if args.debug:
        logger.setLevel(logging.DEBUG)
        formatter = logging.Formatter(
            "%(asctime)s %(name)s (%(lineno)d) %(levelname)s: %(message)s",
            datefmt="[%Y-%m-%d %H:%M:%S]",
        )
        for handler in logger.handlers:
            handler.setFormatter(formatter)
    try:
        PgSummary(
            host=args.host,
            database=args.database,
            user=args.user,
            table_or_view=args.table,
            port=args.port,
            schema=args.schema,
            outfile=args.outfile,
            include_src=args.include_src,
        ).summary()
    except SystemExit as x:
        sys.exit(x.code)
    except KeyboardInterrupt:
        logger.error("Script cancelled by user.")
        sys.exit(1)
    except psycopg2.Error:
        strace = traceback.extract_tb(sys.exc_info()[2])[-1:]
        lno = strace[0][1]
        src = strace[0][3]
        logger.error(
            f"{sys.exc_info()[1]}",
        )
        sys.exit(1)
    except ValueError:
        strace = traceback.extract_tb(sys.exc_info()[2])[-1:]
        lno = strace[0][1]
        src = strace[0][3]
        logger.error(
            f"ValueError on line {lno}: {sys.exc_info()[1]}",
        )
        sys.exit(1)
    except Exception:
        strace = traceback.extract_tb(sys.exc_info()[2])[-1:]
        lno = strace[0][1]
        src = strace[0][3]
        logger.error(
            f"Uncaught exception {sys.exc_info()[0]!s} ({sys.exc_info()[1]}) on line {lno} ({src}).",
        )
        sys.exit(1)


if __name__ == "__main__":
    cli()
