from .asmbl_finder import find_all_assemblies
from .locus_finder import extract_prefixes, retrieve_genes_references
from .coverage_analyser import screen_cov_data
from .mutant_finder import screen_mut_data

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
import textwrap as tw
import openpyxl, os


# Constants
COV_PATTERN = str('coverage_denovo_genes_')
COV_EXT = str('.txt')
FA_PATTERN = str('genes_found_denovo_assembly_')
FA_EXT = str('.fa')
XLS_ROW_LIMIT = 1048576


def _create_new_sheet(WORKBOOK: Workbook, SHEET_NAME: str) -> Worksheet:
    """
        Create a new sheet in the current workbook. Typically called when
        the number of rows written reach XLS_ROW_LIMIT. This new sheet
        replaces the current sheet in collate_results(), so it can continue
        writing data.
    """
    # Create new sheet
    sheet_digits = list()
    is_cont = True if SHEET_NAME.count('(') > 0 else False
    if is_cont:
        [sheet_digits.append(i) for i in SHEET_NAME.split('(')[-1] if i.isdigit()]
        if len(sheet_digits) == 0:  # Continuation exists, but only one
            titlename = SHEET_NAME.replace(')', ' 2)')
        else:  # Continuation exists, multiple times
            sheet_number = str(int(str('').join(sheet_digits)) + 1)  # Update continuation
            titlename = SHEET_NAME.split('(')[0] + str('(') + sheet_number + str(')')
    else:
        titlename = SHEET_NAME + str(' (Cont)')
    sheet = WORKBOOK.create_sheet(title=titlename)
    # Add headers to new sheet
    headers = list(['Sequence ID', 'Query Prot. (% ID)', 'Ref. species',
                    'Gene copy number', 'Other', 'Mutations found',
                    'Assembly location'])
    sheet['A1'] = headers[0]
    sheet['B1'] = headers[1]
    sheet['C1'] = headers[2]
    sheet['D1'] = headers[3]
    sheet['E1'] = headers[4]
    sheet['F1'] = headers[5]
    sheet['G1'] = headers[6]
    sheet['A1'].font = openpyxl.styles.Font(bold=True)
    sheet['B1'].font = openpyxl.styles.Font(bold=True)
    sheet['C1'].font = openpyxl.styles.Font(bold=True)
    sheet['D1'].font = openpyxl.styles.Font(bold=True)
    sheet['E1'].font = openpyxl.styles.Font(bold=True)
    sheet['F1'].font = openpyxl.styles.Font(bold=True)
    sheet['G1'].font = openpyxl.styles.Font(bold=True)
    # Re-set row counter
    rowN = 2
    return sheet, rowN


def collate_results(PATH: str, GENES_PATH: str, SPREADSHEET: str,
                    FNAME_PREFIX: str, DE_NOVO: bool) -> str:
    """
        Screen PATH for Hound output files, and organise results into
        a human-readable spreadsheet. GENES_PATH contains the path to
        the file(s) containing the query sequenced used during the analysis
        with Hound.
    """
    global FA_EXT
    global COV_EXT

    if DE_NOVO is True:
        ASMBL_DIR = str('assemblies/de_novo/')
    else:
        ASMBL_DIR = str('assemblies/reference_mapped/')

    if FNAME_PREFIX is not None:
        FA_EXT = str('_') + FNAME_PREFIX + FA_EXT
        COV_EXT = str('_') + FNAME_PREFIX + COV_EXT

    if PATH[-1] != str('/'):
        PATH = PATH + str('/')

    if GENES_PATH[-1] != str('/') and\
       GENES_PATH.split('/')[-1].find('.fa') == -1:
        GENES_PATH = GENES_PATH + str('/')

    if os.path.exists(PATH + str('reads/')) is True:
        DIR_LIST = list(['', ])  # Empty list, but of type str
    else:
        DIR_LIST = sorted([DIR for DIR in os.listdir(PATH) if str('.') not in DIR])  # openpyxl cannot sort spreadsheets, do it here
        if os.path.exists(PATH + DIR_LIST[0] + str('/reads/')) is True:
            DIR_LIST = list(['', ])  # Empty list, but of type str

    wb = openpyxl.Workbook()  # Create spreadsheet
    headers = list(['Sequence ID', 'Query Prot. (% ID)', 'Ref. species',
                    'Gene copy number', 'Other', 'Mutations found',
                    'Assembly location'])
    for DIR in DIR_LIST:
        # Format workbook
        if len(DIR) == 0:
            sheet = wb.create_sheet(title=PATH.split('/')[-2])  # Add new workbook
        else:
            sheet = wb.create_sheet(title=DIR)  # Add new workbook
        sheet['A1'] = headers[0]
        sheet['B1'] = headers[1]
        sheet['C1'] = headers[2]
        sheet['D1'] = headers[3]
        sheet['E1'] = headers[4]
        sheet['F1'] = headers[5]
        sheet['G1'] = headers[6]
        sheet['A1'].font = openpyxl.styles.Font(bold=True)
        sheet['B1'].font = openpyxl.styles.Font(bold=True)
        sheet['C1'].font = openpyxl.styles.Font(bold=True)
        sheet['D1'].font = openpyxl.styles.Font(bold=True)
        sheet['E1'].font = openpyxl.styles.Font(bold=True)
        sheet['F1'].font = openpyxl.styles.Font(bold=True)
        sheet['G1'].font = openpyxl.styles.Font(bold=True)

        # Fill spreadsheet
        if len(DIR) > 0:
            DIR = DIR + str('/')
        TOTAL_ASMBL = sorted(find_all_assemblies(PATH + DIR + ASMBL_DIR))  # openpyxl cannot sort spreadsheets, do it here
        PREFIX_LIST = sorted(extract_prefixes(PATH, DIR, FA_PATTERN, FA_EXT))  # openpyxl cannot sort spreadsheets, do it here
        PREFIX_REFERENCES = retrieve_genes_references(PREFIX_LIST, GENES_PATH)

        # If GENES_PATH is a file, process only the PREFIX related to the file
        if GENES_PATH.split('/')[-1].find('.fa') > -1:
            PREFIX_LIST = PREFIX_REFERENCES.keys()

        rowN = 2  # To track rows written.
        if len(TOTAL_ASMBL) > 0:
            for PREFIX in PREFIX_LIST:
                if os.path.exists(PATH + DIR + FA_PATTERN + PREFIX + FA_EXT) is True:
                    # 1)
                    if str('promoter') in PREFIX.lower():
                        print("Processing promoter mutations...")
                        # 2) extract copy number + filename. [DONE]
                        COV_DATA = screen_cov_data(PATH + DIR + COV_PATTERN +
                                                   PREFIX + COV_EXT,
                                                   PREFIX_REFERENCES[PREFIX.removesuffix('_Promoter').split('_')[0]])
                        # 3) extract gene name + gene identity, mutations -> result of mutations (synonymous, non-synonymous, early terminaltion)
                        try:
                            MUT_DATA, TRUNC_DATA, ID_DATA, \
                                INS_DATA, STOP_DATA, \
                                GENE_NAME, \
                                SPECIES_INFO = screen_mut_data(PATH + DIR + FA_PATTERN +
                                                               PREFIX + FA_EXT,
                                                               PREFIX_REFERENCES[PREFIX.removesuffix('_Promoter').split('_')[0]],
                                                               PREFIX_REFERENCES[PREFIX])
                        except KeyError as err:
                            err_msg = str("PREFIX ") + str(err) + str(" must be " \
                                          "contained within the name of the " \
                                          "query FASTA file.")
                            raise KeyError(tw.fill(err_msg))
                    else:
                        # 2) extract copy number + filename. [DONE]
                        COV_DATA = screen_cov_data(PATH + DIR + COV_PATTERN +
                                                   PREFIX + COV_EXT,
                                                   PREFIX_REFERENCES[PREFIX.split('_')[0]])
                        # 3) extract gene name + gene identity, mutations -> result of mutations (synonymous, non-synonymous, early terminaltion)
                        try:
                            MUT_DATA, TRUNC_DATA, ID_DATA, \
                                INS_DATA, STOP_DATA, \
                                GENE_NAME, \
                                SPECIES_INFO = screen_mut_data(PATH + DIR + FA_PATTERN +
                                                               PREFIX + FA_EXT,
                                                               PREFIX_REFERENCES[PREFIX.split('_')[0]])
                        except KeyError as err:
                            err_msg = str("PREFIX ") + str(err) + str(" must be " \
                                          "contained within the name of the " \
                                          "query FASTA file.")
                            raise KeyError(tw.fill(err_msg))

                    for ASMBL in TOTAL_ASMBL:
                        if ASMBL in ID_DATA.keys():
                            for N, gene_info in enumerate(ID_DATA[ASMBL]):
                                GENE_NAME = gene_info[0]
                                if str('promoter') in PREFIX.lower():
                                    GENE_NAME = GENE_NAME + str(' Promoter')
                                IDENTITY = gene_info[1]

                                SPECIES = SPECIES_INFO[ASMBL][0][1]

                                sheet.cell(row=rowN, column=1).value = ASMBL
                                sheet.cell(row=rowN, column=2).value = GENE_NAME
                                sheet.cell(row=rowN, column=3).value = SPECIES

                                if IDENTITY is None:
                                    sheet.cell(row=rowN, column=5).value = str('Gene not found / Identity too restrictive.')
                                else:
                                    if GENE_NAME is False:  # Gene name == PREFIX
                                        sheet.cell(row=rowN, column=2).value = PREFIX.replace('_', ' ') +\
                                                                               str(' (') + IDENTITY + str('%)')
                                    else:  # Gene name in .fasta file
                                        sheet.cell(row=rowN, column=2).value = GENE_NAME +\
                                                                               str(' (') + IDENTITY + str('%)')

                                    OTHER_INF = '' if TRUNC_DATA[ASMBL][N][1] is False else str(round(TRUNC_DATA[ASMBL][N][1]*100, ndigits=1)) + '% deletion'
                                    OTHER_INF += '' if INS_DATA[ASMBL][N][1] is False else '; ' + str(INS_DATA[ASMBL][N][1]) + 'AA insertion'
                                    OTHER_INF += '' if STOP_DATA[ASMBL][N][1] is False else '; NON-SENSE mutation'
                                    if len(COV_DATA) == 0:
                                        # No coverage computed
                                        sheet.cell(row=rowN, column=4).value = str('Data not requested')
                                    elif ASMBL in COV_DATA:
                                        # Coverage data computed succesfully
                                        coverage = [e for e in COV_DATA[ASMBL] if GENE_NAME.removesuffix('_Promoter') == e[0]]
                                        if len(coverage) == 0:
                                            # Coverage data requested but not computed
                                            OTHER_INF += '; Coverage FAILED'
                                        else:
                                            coverage = coverage[0]
                                            sheet.cell(row=rowN, column=4).value = coverage[2]  # 0=gene name, 1=contig, 2=rel coverage
                                    else:
                                        # Coverage data requested but not computed
                                        OTHER_INF += '; Coverage FAILED'

                                    sheet.cell(row=rowN, column=5).value = OTHER_INF

                                    if MUT_DATA[ASMBL][N][1] is None:
                                        sheet.cell(row=rowN, column=6).value = str('None')
                                    else:
                                        if len(MUT_DATA[ASMBL][N][1]) == 2 and type(MUT_DATA[ASMBL][N][1][0]) is not list:  # Only one mutation
                                            mut = MUT_DATA[ASMBL][N][1]
                                            sheet.cell(row=rowN, column=6).value = str(int(mut[0])+1) + mut[1]  # workaround mutation at position 0
                                        else:  # Multiple mutations
                                            mut_list = list()
                                            for mut in MUT_DATA[ASMBL][N][1]:
                                                mut_list.append(str(int(mut[0])+1) + mut[1])  # workaround mutation at position 0
                                            sheet.cell(row=rowN, column=6).value = str('; ').join(mut_list)

                                # Update row number in FASTA with multiple queries
                                if len(ID_DATA[ASMBL]) > 1 and rowN == XLS_ROW_LIMIT:
                                    sheet, rowN = _create_new_sheet(wb, sheet.title)
                                elif len(ID_DATA[ASMBL]) > 1:
                                    rowN += 1
                        else:
                            sheet.cell(row=rowN, column=1).value = ASMBL
                            # sheet.cell(row=rowN, column=3).value = SPECIES
                            if str('promoter') in PREFIX.lower():
                                GENE_NAME = GENE_NAME + str(' Promoter')
                            sheet.cell(row=rowN, column=2).value = GENE_NAME
                            sheet.cell(row=rowN, column=5).value = str('Gene not found / Identity too restrictive.')

                        sheet.cell(row=rowN, column=7).value = PATH + DIR
                        # Update row number
                        if rowN == XLS_ROW_LIMIT:
                            sheet, rowN = _create_new_sheet(wb, sheet.title)
                        else:
                            rowN += 1
        else:
            print("No assemblies found, nothing to analyse.")

    # Save summary
    del wb['Sheet']  # Delete default spreadsheet
    wb.save(SPREADSHEET)
    wb.close()
    return


