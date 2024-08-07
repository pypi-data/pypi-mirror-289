## 
# Copyright (c) 2023 Chongqing Spiritlong Technology Co., Ltd.
# All rights reserved.
# 
# @author	arthuryang
# @brief	excel工具集，只支持xlsx。
##

import	openpyxl
import	re
import	os

## 打开xlsx文件
def open_xlsx(filename):
	try:
		book	= openpyxl.load_workbook(filename)
		return book
	except Exception as ex:
		print(str(ex))
		exit

## 加载excel中的数据，可以指定标题行
#	sheet		表格sheet对象
#	title_row	标题行所在行，从1开始
# {}
def get_data_with_title_row(sheet, title_row=1):
	data	= []
	titles	= []
	# 获得标题
	for column in range(1, sheet.max_column+1):
		titles.append(cell_string(sheet, title_row, column))
	
	# 获取数据
	for row in range(title_row+1, sheet.max_row+1):
		record	= {}
		for column in range(1, sheet.max_column+1):
			record[titles[column-1]]	= cell_string(sheet, row, column)
		data.append(record)
	
	return data

## excel读取出None的情况要转换成空字符串
def cell_string(sheet, i, j):
	value	= sheet.cell(i, j).value
	if value is None:
		value	= ""
	return str(value)

## 填充excel单元格，可以指定格式
def cell_fill(sheet, i, j, value, style={
					'font'		: None,
					'fill'		: None, 
					'border'	: None, 
					'alignment'	: None, 
					'number_format'	: None,
				}):
	cell		= sheet.cell(row=i, column=j)
	cell.value	= value
	
	if 'font' in style and style['font']:
		cell.font		= style['font']
	if 'fill' in style and style['fill']:
		cell.fill		= style['fill']
	if 'border' in style and style['border']:
		cell.border		= style['border']
	if 'alignment' in style and style['alignment']:
		cell.alignment		= style['alignment']
	if 'number_format' in style and style['number_format']:
		cell.number_format	= style['number_format']

## 自动调整指定列的宽度
#	column		列（从1开始）
#	width		0自动
#	max_width	最大宽度限制
def adjust_column_width(sheet, column, width=0, max_width=100):
	# 自动计算最大宽度
	if width==0:
		width	= 1
		for row in range(1, sheet.max_row+1):
			# 对于中文字符和非中文字符单独计算宽度
			s	= cell_string(sheet, row, column)
			w	= 0
			for c in s:
				w	+= (1 if ord(c)<128 else 2)
			width	= max(width, w+1)

	# 列宽得有最大限制
	if width>max_width:
		width	= max_width

	sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width	= width

## 自动调整所有列的宽度
def adjust_all_column_width(sheet, max_width=100):
	for column in range(1, sheet.max_column+1):
		adjust_column_width(sheet, column, max_width=max_width)

## 将'A1'转换为（1,1)
def cell_coordinate(cell_string):
	column_letters, row	= openpyxl.utils.cell.coordinate_from_string(cell_string)
	column			= openpyxl.utils.cell.column_index_from_string(column_letters)
	return (row, column)

## 将（1，1）转换为'A1'
def cell_code(row, column):
	return f"{openpyxl.utils.cell.get_column_letter(column)}{row}"

# excel单元格格式：Calibri字体，11号，黑色
font_default	= openpyxl.styles.Font(
	name		= 'Calibri',
	size		= 11,
	bold		= False,
	vertAlign	= None,
	underline	= 'none',
	strike		= False,
	color		= 'FF000000',
)

# excel单元格格式：Calibri字体，11号，白色
font_title	= openpyxl.styles.Font(
	name		= 'Calibri',
	size		= 11,
	bold		= False,
	vertAlign	= None,
	underline	= 'none',
	strike		= False,
	color		= 'FFFFFFFF',
)

# excel单元格格式：Calibri字体，11号，红色
font_red	= openpyxl.styles.Font(
	name		= 'Calibri',
	size		= 11,
	bold		= False,
	vertAlign	= None,
	underline	= 'none',
	strike		= False,
	color		= 'FFFF0000',
)

# excel单元格标题格式：蓝色填充
fill_title	= openpyxl.styles.PatternFill(
	fill_type	= 'solid',
	fgColor		= 'FF4169E1',
)

# excel单元格格式：居中对齐
alignment_center	= openpyxl.styles.Alignment(
	horizontal	= 'center',
	vertical	= 'center',
	wrap_text	= False,
)

# excel单元格格式：左对齐
alignment_left	= openpyxl.styles.Alignment(
	horizontal	= 'left',
	vertical	= 'center',
	wrap_text	= False,
)

# excel单元格格式：右对齐
alignment_right	= openpyxl.styles.Alignment(
	horizontal	= 'right',
	vertical	= 'center',
	wrap_text	= False,
)

################ 单元格样式：字体 ################
# 默认黑色11号Calibri
def font(name='Calibri', size=11, color='FF000000', bold=False):
	return openpyxl.styles.Font(
		name		= name,
		size		= size,
		bold		= bold,
		vertAlign	= None,
		underline	= 'none',
		strike		= False,
		color		= color,
	)

################ 单元格样式：填充 ################
# 默认白色
def fill_with_color(color='FFFFFFFF'):
	return openpyxl.styles.PatternFill(
		fill_type	= 'solid',
		fgColor		= color,
	)

################ 单元格样式：对齐 ################
# 居中对齐
alignment_center	= openpyxl.styles.Alignment(
	horizontal	= 'center',
	vertical	= 'center',
	wrap_text	= False,
)

# 左对齐
alignment_left	= openpyxl.styles.Alignment(
	horizontal	= 'left',
	vertical	= 'center',
	wrap_text	= False,
)

# 右对齐
alignment_right	= openpyxl.styles.Alignment(
	horizontal	= 'right',
	vertical	= 'center',
	wrap_text	= False,
)

################ 单元格样式 ################
style_title	= {
	'font'		: font_title,
	'fill'		: fill_title,
	'alignment'	: alignment_center, 
}

style_text	= {
	'font'		: font_default,
	'fill'		: None,
	'alignment'	: alignment_left, 
}

## 保存记录到excel文件
#	str 		filename 	文件名,带路径
#	[{}]		records		记录数据，每个记录是一个列表
#	[]/{}		titles		表头部,{列名:key}或[列名]，None表示使用所有记录中依次出现过的所有名称
#	str		sheet_name	sheet名称
# 	{}		style_title	标题样式,{font,fill,border,alignment,number_format}
# 	{}		style_text	内容样式,{font,fill,border,alignment,number_format}	
def records_to_excel(filename, records, titles=None, sheet_name=None, style_title=style_title, style_text=style_text):
	# 标题
	if not titles:
		# 未提供则使用数据中出现了的字段名
		titles	= {k:k for r in records for k in r}
	elif isinstance(titles, list):
		# []则表示字段名和标题名相同
		titles	= {t:t for t in titles}
	elif not isinstance(titles, dict):
		# 此时必须是字典了
		print(f"必须为标题列表、字典或None")
		return
	
	# 创建sheet
	book 		= openpyxl.Workbook()
	sheet		= book.active
	if sheet_name:
		sheet.title 	= sheet_name

	# 写入标题
	for i, t in enumerate(titles):
		cell_fill(sheet, 1, i+1, t, style_title)
	
	# 写入内容
	for i, r in enumerate(records):
		for j, field in enumerate(titles):
			if field in r.keys():
				cell_fill(sheet, i+2, j+1, r[field], style_text)	

	# 自动调整列宽
	adjust_all_column_width(sheet)
	
	# 保存文件，确保是xlsx后缀名
	if not filename.endswith(".xlsx"):
		filename	+= ".xlsx"
	book.save(filename)

## 创建一个文件模板
#	str 		filename 	文件名,带路径
#	[]		columns		[列名]
#	list		list_data	数据
#	str		sheet_name	sheet名称
# 	{}		style_title	标题样式,{font,fill,border,alignment,number_format}
# str
def create_template(filename, columns=['ID'], sheet_name="Sheet1", style_title=style_title):
	# 不存在则创建
	if not os.path.exists(filename):
		book	= openpyxl.Workbook()
		
		# sheet改名
		sheet		= book.active
		sheet.title	= sheet_name

		# 写入标题
		for i, t in enumerate(columns):
			cell_fill(sheet, 1, i+1, t, style	= style_title)

		# 自动调整列宽
		adjust_all_column_width(sheet)

		# 保存文件
		if filename.endswith(".xlsx") == False:
			filename	+= ".xlsx"
		book.save(filename)

# 调试/测试代码
if __name__ == '__main__':
	pass