# -*- coding: utf-8 -*-
"""
Created on 11Dec2023
@author: YueXu

Updated on 01Aug2024
@author: YueXu
1.subitem 填充逻辑更新：更新为填充metadata中Test Name/Timepoint的内容
2.header更新：P/F_clean更新为P/F_clean
3.必填更新：支持numeric的字段
"""

import openpyxl
from  openpyxl.styles import Font,Alignment,Border,Side,PatternFill
import pandas as pd
import datetime
import re
import pdfplumber
import os


class TestingCaseCreator():
    def __init__(self):
        pass
    
    @staticmethod
    def Creator(acrf_path, meta_path,out_path,projectName = None,bcrf_path = None):
        acrf = None
        try:
            annotedcrf_path=acrf_path
            metadata_path=meta_path
            if projectName:
                Debug_path = os.path.join(out_path, projectName + "_" +"output.txt")
            else:
                Debug_path = os.path.join(out_path, "output.txt")
            Debug_path = out_path + "output.txt"
            acrf=pdfplumber.open(annotedcrf_path)

            #---------------------------------------------------------------------------------------------------------------------
            
            f_item=pd.read_excel(metadata_path,sheet_name="Items")
            f_codelist=pd.read_excel(metadata_path,sheet_name="Codelist")
            f_unique=pd.read_excel(metadata_path,sheet_name="UniqueFormCheckList")
            f_cross=pd.read_excel(metadata_path,sheet_name="CrossFormCheckList")
            f_automation=pd.read_excel(metadata_path,sheet_name="AutomationEvents")
            f_related=pd.read_excel(metadata_path,sheet_name="RelatedFormSettings")
            s_visit=pd.concat([f_unique["Visit/Form"],f_cross["Visit Name"],f_automation["Visit Name"]],axis=0).reset_index(drop=True)
            s_form=pd.concat([f_unique["Unique Form"],f_cross["Form Name"],f_automation["Form Name"]],axis=0).reset_index(drop=True)
            s_variable=pd.concat([f_unique["Variable"],f_cross["Variable"]],axis=0).reset_index(drop=True)
            s_question=pd.concat([f_unique["Question"],f_cross["Question"]],axis=0).reset_index(drop=True)
            s_tip=pd.concat([f_unique["Rule Tip"],f_cross["Rule Tip"]],axis=0).reset_index(drop=True)
            s_class=pd.concat([f_unique["Type class"],f_cross["Type class"],f_automation["Type"]],axis=0).reset_index(drop=True)
            s_details=pd.concat([f_unique["Details"],f_cross["Details"],f_automation["Details"]],axis=0).reset_index(drop=True)
            s_status=pd.concat([f_unique["Edit Check Status"],f_cross["Edit Check Status"],f_automation["Status"]],axis=0).reset_index(drop=True)
            s_rule=pd.concat([f_unique["OID "],f_cross["OID "],f_automation["OID "]],axis=0).reset_index(drop=True)
            s_itemoid=pd.concat([f_unique["Item OID"],f_cross["Item OID"]],axis=0).reset_index(drop=True)
            s_itemtype=pd.concat([f_unique["Type"],f_cross["Type"]],axis=0).reset_index(drop=True)
            s_caption=pd.concat([f_unique["Caption"],f_cross["Caption"]],axis=0).reset_index(drop=True)
            s_subitem=pd.concat([f_unique["Test Name/Timepoint"],f_cross["Test Name/Timepoint"]],axis=0).reset_index(drop=True)
            #获取metadata中unique\cross\automation工作簿指定列，并合并成series
            #对于unique或者cross但是此表单没有其它相同变量名的，以item oid来定位item
            #对于corss但是此表单有相同变量名的，无法通过item oid定位item
            #新增的列itemoid itemtype 就是为了判断这种情况
            #新增caption列是为了dependency函数，部分取不到caption的cross EC而新增的列
            
            f_key=["#","Visit # (Tab)","Form","Variable","Item Name (Data Point)","Sub - Item Name (Sub - Data Point)","logic","Query Text","Type Class","Edit Check Status","Comment","Condition (Standard, Tested, New,Tip Update)","Referenced study ID and check #" ,"Test Study ID","Test Subject(All visits have been tested)_clean","Test Value_clean","Query Fired(Yes, No, Enable, Disable, user-defined)_clean","P/F_clean",
            "Test Study ID","Test Subject(All visits have been tested)","Test Value","Query Fired(Yes, No, Enable, Disable, user-defined)","P/F","Tested by/Date","Retest Subject/Visit/Form","Retest Value","Query Fired(Yes, No)","P/F","Retested By/Date","caption","item type","rule type","details","rule oid","Item OID"]
            
            f=pd.DataFrame(index=None,columns=f_key)
            
            f[f_key[1]]=s_visit
            f[f_key[2]]=s_form
            f[f_key[3]]=s_variable
            f[f_key[4]]=s_question
            f[f_key[5]]=s_subitem
            f[f_key[7]]=s_tip
            f[f_key[8]]=s_class
            f[f_key[9]]=s_status
            f[f_key[-6]]=s_caption
            f[f_key[-5]]=s_itemtype
            f[f_key[-3]]=s_details
            f[f_key[-2]]=s_rule
            f[f_key[-1]]=s_itemoid
            f[f_key[11]].fillna(value="New",inplace=True)
            
            #形成文件填充相应列,考虑到testing case 的key可能后期会修改，代码全部用数字来定位列
            #11列全部填充New
            
            f.iloc[0:len(f_unique),-4]="unique"
            f.iloc[len(f_unique):(len(f_unique)+len(f_cross)),-4]="cross"
            #根据长度填充ruletype,用于判断是cross还是unique，cross的值不能在metadata找
            
            #---------------------------------------------------------------------------------------------------------------------
            
            #subitem 列填充
            #如果subitem 列是空值，则填充此字段的caption
            
            def func_caption (x):
                
                    if isinstance(x[f_key[5]],str)!=True:
                        x[f_key[5]]=x[f_key[-6]]
                    return None                                                                                              
            f.apply(func_caption,axis=1)
            
            with open(Debug_path, 'a') as debug:
                print("----------------------caption填充结束-------------------------------", file=debug)
            #---------------------------------------------------------------------------------------------------------------------
            
            def func_visit (x):
                if isinstance(x[f_key[2]],str):
                    f.at[x.name,f_key[1]]=x[f_key[1]].replace("/"+x[f_key[2]],"").replace(",","")
                    return x
            
            f.apply(func_visit,axis=1)
            #删除掉visit列中的表单名称
            with open(Debug_path, 'a') as debug:
                print("----------------------访视列规范化结束-------------------------------", file=debug)
            #---------------------------------------------------------------------------------------------------------------------
            def func_required (x):
               try: 
                   if x[f_key[8]]=="Required Field" :
                          f.at[x.name,f_key[20]]="[blank]"
                          #dirty值为[blank]
                          itemtype=f_item.loc[f_item["OID"]==x[f_key[-1]]].reset_index(drop=True).loc[0,"Item Type"]
                          #通过item oid在item工作簿定位到此字段所在行，重新设置index，取第一行的item type
                          #由于部分字段修改变量名、codelist之后会出现两行，第一行为修改后的，第二行为修改前的，取index为0的值可以确保能取到最新的值
                          if  itemtype=="Radio" or itemtype=="ListBox":
                               codelist=f_item.loc[f_item["OID"]==x[f_key[-1]]].reset_index(drop=True).loc[0,"Code List "]
                               codelist_text=f_codelist.loc[f_codelist["Codelist Name"]==codelist].reset_index(drop=True).loc[0,"Codelist Item Text"]
                               f.at[x.name,f_key[15]]="\""+codelist_text+"\""+" is selected"
                          #如果是radio或者listbox，根据codist到codelist工作簿寻找对应行得到df，重新设置index，统一取index为0的作为测试值
                          elif itemtype=="Optional":
                               itemcaption=f_item.loc[f_item["OID"]==x[f_key[-1]]].reset_index(drop=True).loc[0,"Item Caption"]
                               f.at[x.name,f_key[15]]="\""+itemcaption+"\""+" is checked"
                          #如果是optinal，根据oid在item工作簿定位到此字段所在行，重新设置index，取第一行的caption作为填充值
                          elif itemtype=="Date" or itemtype=="Time" :
                              itemformat=f_item.loc[f_item["OID"]==x[f_key[-1]]].reset_index(drop=True).loc[0,"Format"]
                              time_str="2023/01/01 00:00:00"
                              dt_time=datetime.datetime.strptime(time_str, "%Y/%m/%d %H:%M:%S")
                              itemformat_list_1=itemformat.replace(":","-").split("-")
                              itemformat_list_2=[]
                
                              for i in itemformat_list_1:
                                  if i=="DD":
                                     i="-%d"
                                  elif i=="MM":
                                     i="-%m"
                                  elif i=="MMM":
                                     i="-%b"
                                  elif i=="yyyy":
                                     i="-%Y"
                                  elif i=="HH":
                                     i=":%H"
                                  elif i=="mm":
                                     i=":%M" 
                                  elif i=="ss":
                                     i=":%S"
                                  else:
                                     i=""
                                  itemformat_list_2.append(i)
                                  
                              if len(itemformat_list_2)==1:
                                  datetype=itemformat_list_2[0]
                              elif len(itemformat_list_2)==2:
                                  datetype=itemformat_list_2[0]+itemformat_list_2[1]
                              elif len(itemformat_list_2)==3:
                                  datetype=itemformat_list_2[0]+itemformat_list_2[1]+itemformat_list_2[2]
                              time_str_converted=  dt_time.strftime(datetype.strip("-").strip(":"))  
                              f.at[x.name,f_key[15]]=time_str_converted
                           
                         #如果是Date或者Time，根据oid在item工作簿定位到此字段所在行，重新设置index，取第一行的Format作为此字段的格式
                         #将format中的：全部替换为-，并根据-进行拆分得到itemformat_list_1
                         #对itemformat_list_1进行遍历，将其转换为datetime中的format格式，并append进空列表itemformat_list_2
                         #itemformat_list_2长度最高为3，将各个元素进行凭借，得到最终的格式
                
                          if itemtype=="Reference" :         
                            f.at[x.name,f_key[15]]="1\n2"
                          #若为Reference,clean:1,2,dirty:[blank]
                          
                          if itemtype=="TextArea":  
                            f.at[x.name,f_key[15]]="test"
                          #若为Text,TextArea clean:test,dirty:[blank]
                          if itemtype=="Text":
                          #若为text，需考虑是related setting，和不是related setting       
                            if len(f_related.loc[(x[f_key[1]]==f_related["Visit Name"]) & (x[f_key[2]]==f_related["Form Name"]) & (x[f_key[3]]==f_related["Item Variable"]) ]) ==0:
                                f.at[x.name,f_key[15]]="test"
                                #不是related form setting
                            else:
                                f.at[x.name,f_key[15]]="SN1\nSN2"
                
                                #是related form setting
                          if itemtype=="numeric":  #number
                               f.at[x.name,f_key[15]]="1"
                          return x
               except:
                   pass
                   # with open(Debug_path, 'a') as debug:
                       # print(x[-3],file=debug)
                       # print(x[-2],file=debug)   
                          
                      
            f.apply(func_required,axis=1)
            
            #填充required EC的clean dirty test value
            #通过变量名，到item工作簿检索item type
            #1.若为radio、listbox则检测codelist名称，并到codelist工作簿检索Codelist Item Text，clean:"第一列" is selected;dirty:[blank]
            #2.若为optinal，clean:赋值"caption" is checked;dirty:[blank]
            #3.若为DATE TIME，clean:根据item工作簿对应Format填充，dirty:[blank]
            #4.若为Reference,clean:1,2,dirty:[blank]
            #5.若为TextArea,clean:test,dirty:[blank]
            #6.若为Text，需判断是否是reference，若是则clean值应该为SN1\nSN2"，若不是则clean值为test
            with open(Debug_path, 'a') as debug:
                print("----------------------required field EC 填充完成----------------------------", file=debug)
            #---------------------------------------------------------------------------------------------------------------------
            
            def func_validation (x):
               try:
                    if x[f_key[8]]=="Validation" :
                        itemtype=f_item.loc[f_item["OID"]==x[f_key[-1]]].reset_index(drop=True).loc[0,"Item Type"]
                        if itemtype=="Date" or itemtype=="Time":
                             itemformat=f_item.loc[f_item["OID"]==x[f_key[-1]]].reset_index(drop=True).loc[0,"Format"]
                             time_str="2023/01/01 00:00:00"
                             dt_time=datetime.datetime.strptime(time_str, "%Y/%m/%d %H:%M:%S")
                             itemformat_list_1=itemformat.replace(":","-").split("-")
                             itemformat_list_2=[]
                    
                             for i in itemformat_list_1:
                                 if i=="DD":
                                   i="-%d"
                                 elif i=="MM":
                                   i="-%m"
                                 elif i=="MMM":
                                   i="-%b"
                                 elif i=="yyyy":
                                   i="-%Y"
                                 elif i=="HH":
                                   i=":%H"
                                 elif i=="mm":
                                   i=":%M" 
                                 elif i=="ss":
                                   i=":%S"
                                 else:
                                   i=""
                                 itemformat_list_2.append(i)
                                   
                             if len(itemformat_list_2)==1:
                                 datetype=itemformat_list_2[0]
                             elif len(itemformat_list_2)==2:
                                 datetype=itemformat_list_2[0]+itemformat_list_2[1]
                             elif len(itemformat_list_2)==3:
                                 datetype=itemformat_list_2[0]+itemformat_list_2[1]+itemformat_list_2[2]
                             time_str_converted=  dt_time.strftime(datetype.strip("-").strip(":"))  
                             if re.search("isValid",x[f_key[-3]]):
                               f.at[x.name,f_key[15]]=time_str_converted
                               f.at[x.name,f_key[20]]=itemformat.strip("-").strip(":").replace("DD","UNK").replace("MMM","UNK").replace("MM","UNK").replace("HH","UNK").replace("mm","UNK").replace("ss","UNK").replace("yyyy","2023")
                               #先replaceMMMM再replaceMM.如果先MM再MMM，遇到MMM后会先将其中的MM替换成UNK，导致最终结果出现UNKM
                             elif re.search("isUNK",x[f_key[-3]]):
                               f.at[x.name,f_key[15]]=itemformat.strip("-").strip(":").replace("DD","UNK").replace("MMM","UNK").replace("MM","UNK").replace("HH","UNK").replace("mm","UNK").replace("ss","UNK").replace("yyyy","2023")
                               f.at[x.name,f_key[20]]=itemformat.strip("-").strip(":").replace("DD","Blank").replace("MMM","Blank").replace("MM","Blank").replace("HH","Blank").replace("mm","Blank").replace("ss","Blank").replace("yyyy","2023")
                    
                             return x
               except:
                    pass
               #     with open(Debug_path, 'a') as debug:
               #         print(x[-2],file=debug) 
                    
            f.apply(func_validation,axis=1)
            #针对date和time的validationEC填充值
            #对于date的dirty值，年份不可UNK
            
            # with open(Debug_path, 'a') as debug:
            #     print("----------------------validation EC 填充完成----------------------------", file=debug)
            #---------------------------------------------------------------------------------------------------------------------
            
            def func_dependency (x):
                try:
                    if x[f_key[8]]=="Dependency Rule" or x[f_key[8]]=="show unique details" or x[f_key[8]]=="show form" or x[f_key[8]]=="show visit":
                      if re.search("=\$",x[f_key[-3]])==None and re.search("\(\$",x[f_key[-3]])==None  and re.search("getSubjectArmCode",x[f_key[-3]])==None and  re.search("\$",x[f_key[-3]])!=None:
                    #判断无括号嵌套,判断显隐触发条件不是必填，判断逻辑符两边一边为变量一边为值
                    #有括号嵌套一定会出现（$...）
                    #显隐触发条件是必填一定会出现tonum($..)todate($..)mustanswer($..)ischecked($..)
                    #以变量==变量或变量!=变量一定会出现$...==$..
                        list_dependency=x[f_key[-3]].replace("||","&&").split("&&")      
                        #list_dependecy为[$*.*.CTCCHG4=="Y",$*.*.CTCCHG4=="Y".....]
                        
                        list_dependency_num=len(list_dependency)
                        #list_dependency_num判断details里组分个数，如A记为1个，A&&B||C记为3个
                        #此步作用是，对于组分只有一个，且是radio，且codlist里有2-3个value(如a,b,c)，在填充dirty值的时候，不应该填充a is not selected，而应该填充b is selected
                
                        global cleandata
                        global dirtydata
                        details_clean=x[f_key[-3]]
                        details_dirty=x[f_key[-3]]
                
                        for i in list_dependency:
                            item_split=i.replace("!","=").split("=")[0].strip("$").split(".",2)
                            #item_split为$*.*.CTCCHG4=="Y"左边部分的拆分list，[*,*,CTCCHG4]
                            #由于会出现$*.*.VSSTAT.6692!="on"，限定拆分2次
                            if item_split[0]=="*" and item_split[1]=="*"  and item_split[2]!="*":
                                if re.search("\.",item_split[2])!=None:
                                #判断是否是变量名加oid这种情况，如$*.*.VSSTAT.6692!="on",当同一表单多个相同变量名时会出现这种情况
                                    if x[f_key[-4]]=="unique":
                                    #如果规则是unique的，则oid也是unique表单的oid，可在metedata文件中找值
                                      item_oid=item_split[2].split(".")[1]
                                      item_oid=f"Item.{item_oid}"
                                      item_type=f_item.loc[f_item["OID"]==item_oid].reset_index(drop=True).loc[0,"Item Type"]
                                      item_question=f_item.loc[f_item["OID"]==item_oid].reset_index(drop=True).loc[0,"Question"]
                                      item_caption=f_item.loc[f_item["OID"]==item_oid].reset_index(drop=True).loc[0,"Item Caption"]
                                      if item_type=="Radio" or item_type=="ListBox" :
                                          codelist_name=f_item.loc[f_item["OID"]==item_oid].reset_index(drop=True).loc[0,"Code List "]
                                    else:
                                    #若不是，则oid是访视表单的oid，无法在metedata中直接定位字段
                                    #可以通过访视表单中的oid在acrf中定位字段，再通过表单名称、caption、变量名称等定位字段                   
                                      item_oid=item_split[2].split(".")[1]
                                      item_variable=item_split[2].split(".")[0]
                                      for page in acrf.pages:
                                        if re.search(f"\n(.*?)Item{item_oid}(.*?)\n").group(0)!=None:
                                            item_caption=re.search(f"\n(.*?)Item{item_oid}(.*?)\n").group(1)
                                            break
                                      item_type=f_item.loc[(f_item["Item Caption"]==item_caption) & (f_item["Form Name"]==x[2]) & (f_item["CDASH Variable"]==item_variable)].reset_index(drop=True).loc[0,"Item Type"]
                                      item_question=f_item.loc[(f_item["Item Caption"]==item_caption) & (f_item["Form Name"]==x[2]) & (f_item["CDASH Variable"]==item_variable)].reset_index(drop=True).loc[0,"Question"]
                                      if item_type=="Radio" or item_type=="ListBox" :
                                          codelist_name=f_item.loc[(f_item["Item Caption"]==item_caption) & (f_item["Form Name"]==x[2]) & (f_item["CDASH Variable"]==item_variable)].reset_index(drop=True).loc[0,"Code List "] 
                                          
                                elif re.search("\.",item_split[2])==None:
                                    item_type=f_item.loc[(f_item["Form Name"]==x[f_key[2]]) & (f_item["CDASH Variable"]==item_split[2])].reset_index(drop=True).loc[0,"Item Type"]
                                    item_question=f_item.loc[(f_item["Form Name"]==x[f_key[2]]) & (f_item["CDASH Variable"]==item_split[2])].reset_index(drop=True).loc[0,"Question"]
                                    item_caption=f_item.loc[(f_item["Form Name"]==x[f_key[2]]) & (f_item["CDASH Variable"]==item_split[2])].reset_index(drop=True).loc[0,"Item Caption"]
                                    if item_type=="Radio" or item_type=="ListBox" :
                                        codelist_name=f_item.loc[(f_item["Form Name"]==x[f_key[2]]) & (f_item["CDASH Variable"]==item_split[2])].reset_index(drop=True).loc[0,"Code List "]              
                                if item_type=="Optional":
                                    if item_question!=item_caption:
                                      if re.search("==",i)!=None:
                                        cleandata=f"\"{item_question}\" selects \"{item_caption}\""
                                        dirtydata=f"\"{item_question}\" does not select \"{item_caption}\""
                                      else:
                                        cleandata=f"\"{item_question}\" does not select \"{item_caption}\""
                                        dirtydata=f"\"{item_question}\" selects \"{item_caption}\""    
                                    else:
                                      if re.search("==",i)!=None:
                                        cleandata=f"\"{item_caption}\" is selected "
                                        dirtydata=f"\"{item_caption}\" is not selected"
                                      else:
                                        cleandata=f"\"{item_caption}\" is not selected"
                                        dirtydata=f"\"{item_caption}\" is selected"
                                        
                                elif item_type=="Radio" or item_type=="ListBox" :
                                  i_replace=i.replace("\'","\"")
                                  #进行测试时发现部分EC用的双引号，部分又用单引号，全部替换为双引号进行匹配
                                  codelist_value=re.search("\"(.*?)\"",i_replace).group().strip("\"")
                                  codelist_text=f_codelist.loc[(f_codelist["Codelist Name"]==codelist_name) & (f_codelist["Codelist Item Value"]==codelist_value)].reset_index(drop=True).loc[0,"Codelist Item Text"]
                                  codelist_num=len(f_codelist.loc[f_codelist["Codelist Name"]==codelist_name])
                                  #codelist_num是某codelist_name下value的个数
                                  if codelist_num==2 and list_dependency_num==1:
                                  #对于仅有一个组分的details，需判断其codelist_name下是否存在2个value，若存在(如a,b)，在填充dirty值的时候，不应该填充a is not selected，而应该填充b is selected
                                    codelist_text_1= f_codelist.loc[(f_codelist["Codelist Name"]==codelist_name) & (f_codelist["Codelist Item Value"]!=codelist_value)].reset_index(drop=True).loc[0,"Codelist Item Text"]
                                    if re.search("==",i_replace)!=None:                            
                                        cleandata=f"\"{item_question}\" selects \"{codelist_text}\""
                                        dirtydata=f"\"{item_question}\" selects \"{codelist_text_1}\""
                                    else:
                                        cleandata=f"\"{item_question}\" selects \"{codelist_text_1}\""
                                        dirtydata=f"\"{item_question}\" selects \"{codelist_text}\""
                                  else:
                                    if re.search("==",i_replace)!=None:
                                        cleandata=f"\"{item_question}\" selects \"{codelist_text}\""
                                        dirtydata=f"\"{item_question}\" does not select \"{codelist_text}\""
                                    else:
                                        cleandata=f"\"{item_question}\" does not select \"{codelist_text}\""
                                        dirtydata=f"\"{item_question}\" selects \"{codelist_text}\""
                
                            elif item_split[1]!="*" and item_split[2]!="*":
                                if re.search("\.",item_split[2])!=None:
                                    if x[f_key[-4]]=="unique":
                                        item_oid=item_split[2].split(".")[1]
                                        item_oid=f"Item.{item_oid}"
                                        item_type=f_item.loc[f_item["OID"]==item_oid].reset_index(drop=True).loc[0,"Item Type"] 
                                        #如果是unique表单中的依赖且是$*.Screening.LBDAT.6287这种形式，可以通过oid定位字段
                                        item_question=f_item.loc[f_item["OID"]==item_oid].reset_index(drop=True).loc[0,"Question"]
                                        item_caption=f_item.loc[f_item["OID"]==item_oid].reset_index(drop=True).loc[0,"Item Caption"]
                                        if item_type=="Radio" or item_type=="ListBox" :
                                            codelist_name=f_item.loc[f_item["OID"]==item_oid].reset_index(drop=True).loc[0,"Code List "]
                                    else:
                                    #若不是，则oid是访视表单的oid，无法在metedata中直接定位字段
                                    #可以通过访视表单中的oid在acrf中定位字段，再通过表单名称、caption、变量名称等定位字段  
                #####这里还需要核实一下cross中form name是不是和unique的是一致的，还是访视中的名称，如果是在这里是只适用unique和访视中名称一致的                  
                                        item_oid=item_split[2].split(".")[1]
                                        item_variable=item_split[2].split(".")[0]
                                        for page in acrf.pages:
                                         if re.search(f"\n(.*?)Item{item_oid}(.*?)\n").group(0)!=None:
                                             item_caption=re.search(f"\n(.*?)Item{item_oid}(.*?)\n").group(1)
                                             break
                                        item_type=f_item.loc[(f_item["Item Caption"]==item_caption) & (f_item["Form Name"]==item_split[1]) & (f_item["CDASH Variable"]==item_variable)].reset_index(drop=True).loc[0,"Item Type"]
                                        item_question=f_item.loc[(f_item["Item Caption"]==item_caption) & (f_item["Form Name"]==item_split[1]) & (f_item["CDASH Variable"]==item_variable)].reset_index(drop=True).loc[0,"Question"]         
                                        if item_type=="Radio" or item_type=="ListBox" :
                                          codelist_name=f_item.loc[(f_item["Item Caption"]==item_caption) & (f_item["Form Name"]==item_split[1]) & (f_item["CDASH Variable"]==item_variable)].reset_index(drop=True).loc[0,"Code List "]                 
                            
                                elif re.search("\.",item_split[2])==None:
                                    item_type=f_item.loc[(f_item["Form Name"]==item_split[1]) & (f_item["CDASH Variable"]==item_split[2])].reset_index(drop=True).loc[0,"Item Type"]
                                    #如果是是$*.Screening.LBDAT这种形式，需要先定位访视再定位字段
                                    item_question=f_item.loc[(f_item["Form Name"]==item_split[1]) & (f_item["CDASH Variable"]==item_split[2])].reset_index(drop=True).loc[0,"Question"]
                                    item_caption=f_item.loc[(f_item["Form Name"]==item_split[1]) & (f_item["CDASH Variable"]==item_split[2])].reset_index(drop=True).loc[0,"Item Caption"]
                                    if item_type=="Radio" or item_type=="ListBox" :
                                      codelist_name=f_item.loc[(f_item["Form Name"]==item_split[1]) & (f_item["CDASH Variable"]==item_split[2])].reset_index(drop=True).loc[0,"Code List "]
                                if item_type=="Optional":
                                  if item_split[0]=="*":
                                    if item_question!=item_caption:
                                      if re.search("==",i)!=None:
                                        cleandata=f"\"{item_question}\" selects \"{item_caption}\" in \"{item_split[1]}\" form"
                                        dirtydata=f"\"{item_question}\" does not select \"{item_caption}\" in \"{item_split[1]}\" form"
                                      else:
                                        cleandata=f"\"{item_question}\" does not select \"{item_caption}\" in \"{item_split[1]}\" form"
                                        dirtydata=f"\"{item_question}\" selects \"{item_caption}\" in \"{item_split[1]}\" form"    
                                    else:
                                      if re.search("==",i)!=None:
                                        cleandata=f"\"{item_caption}\" is selected in \"{item_split[1]}\" form"
                                        dirtydata=f"\"{item_caption}\" is not selected in \"{item_split[1]}\" form"
                                      else:
                                        cleandata=f"\"{item_caption}\" is not selected in \"{item_split[1]}\" form"
                                        dirtydata=f"\"{item_caption}\" is selected in \"{item_split[1]}\" form"
                                  elif item_split[0]!="*":
                                    if item_question!=item_caption:
                                      if re.search("==",i)!=None:
                                        cleandata=f"\"{item_question}\" selects \"{item_caption}\" in {item_split[0]} \"{item_split[1]}\" form"
                                        dirtydata=f"\"{item_question}\" does not select \"{item_caption}\" in {item_split[0]} \"{item_split[1]}\" form"
                                      else:
                                        cleandata=f"\"{item_question}\" does not select \"{item_caption}\" in {item_split[0]} \"{item_split[1]}\" form"
                                        dirtydata=f"\"{item_question}\" selects \"{item_caption}\" in {item_split[0]} \"{item_split[1]}\" form"    
                                    else:
                                      if re.search("==",i)!=None:
                                        cleandata=f"\"{item_caption}\" is selected in {item_split[0]} \"{item_split[1]}\" form"
                                        dirtydata=f"\"{item_caption}\" is not selected in {item_split[0]} \"{item_split[1]}\" form"
                                      else:
                                        cleandata=f"\"{item_caption}\" is not selected in {item_split[0]} \"{item_split[1]}\" form"
                                        dirtydata=f"\"{item_caption}\" is selected in {item_split[0]} \"{item_split[1]}\" form"                      
                                      
                                      
                                elif item_type=="Radio" or item_type=="ListBox" :
                                  i_replace=i.replace("\'","\"")
                                  #进行测试时发现部分EC用的双引号，部分又用单引号，全部替换为双引号进行匹配
                                  codelist_value=re.search("\"(.*?)\"",i_replace).group().strip("\"")
                                  try:
                                    codelist_text=f_codelist.loc[(f_codelist["Codelist Name"]==codelist_name) & (f_codelist["Codelist Item Value"]==codelist_value)].reset_index(drop=True).loc[0,"Codelist Item Text"]
                                  except:
                                    print(x[-2])
                                    return None
                                  codelist_num=len(f_codelist.loc[f_codelist["Codelist Name"]==codelist_name])
                                  #codelist_num是某codelist_name下value的个数
                                  if item_split[0]=="*":
                                      if codelist_num==2 and list_dependency_num==1:
                                      #对于仅有一个组分的details，需判断其codelist_name下是否存在2个value，若存在(如a,b)，在填充dirty值的时候，不应该填充a is not selected，而应该填充b is selected
                                        codelist_text_1= f_codelist.loc[(f_codelist["Codelist Name"]==codelist_name) & (f_codelist["Codelist Item Value"]!=codelist_value)].reset_index(drop=True).loc[0,"Codelist Item Text"]
                                        if re.search("==",i_replace)!=None:                            
                                            cleandata=f"\"{item_question}\" selects \"{codelist_text}\" in {item_split[0]} \"{item_split[1]}\" form"
                                            dirtydata=f"\"{item_question}\" selects \"{codelist_text_1}\" in {item_split[0]} \"{item_split[1]}\" form"
                                        else:
                                            cleandata=f"\"{item_question}\" selects \"{codelist_text_1}\" in {item_split[0]} \"{item_split[1]}\" form"
                                            dirtydata=f"\"{item_question}\" selects \"{codelist_text}\" in {item_split[0]} \"{item_split[1]}\" form"    
                                      else:                      
                                        if re.search("==",i_replace)!=None:
                                            cleandata=f"\"{item_question}\" selects \"{codelist_text}\" in {item_split[0]} \"{item_split[1]}\" form"
                                            dirtydata=f"\"{item_question}\" does not select \"{codelist_text}\" in {item_split[0]} \"{item_split[1]}\" form"
                                        else:
                                            cleandata=f"\"{item_question}\" does not select \"{codelist_text}\" in {item_split[0]} \"{item_split[1]}\" form"
                                            dirtydata=f"\"{item_question}\" selects \"{codelist_text}\" in {item_split[0]} \"{item_split[1]}\" form" 
                                  elif item_split[0]!="*":   
                                      if codelist_num==2 and list_dependency_num==1:
                                      #对于仅有一个组分的details，需判断其codelist_name下是否存在2个value，若存在(如a,b)，在填充dirty值的时候，不应该填充a is not selected，而应该填充b is selected
                                        codelist_text_1= f_codelist.loc[(f_codelist["Codelist Name"]==codelist_name) & (f_codelist["Codelist Item Value"]!=codelist_value)].reset_index(drop=True).loc[0,"Codelist Item Text"]
                                        if re.search("==",i_replace)!=None:                            
                                            cleandata=f"\"{item_question}\" selects \"{codelist_text}\" in {item_split[0]} \"{item_split[1]}\" form"
                                            dirtydata=f"\"{item_question}\" selects \"{codelist_text_1}\" in {item_split[0]} \"{item_split[1]}\" form"
                                        else:
                                            cleandata=f"\"{item_question}\" selects \"{codelist_text_1}\" in {item_split[0]} \"{item_split[1]}\" form"
                                            dirtydata=f"\"{item_question}\" selects \"{codelist_text}\" in {item_split[0]} \"{item_split[1]}\" form"    
                                      else:                      
                                        if re.search("==",i_replace)!=None:
                                            cleandata=f"\"{item_question}\" selects \"{codelist_text}\" in {item_split[0]} \"{item_split[1]}\" form"
                                            dirtydata=f"\"{item_question}\" does not select \"{codelist_text}\" in {item_split[0]} \"{item_split[1]}\" form"
                                        else:
                                            cleandata=f"\"{item_question}\" does not select \"{codelist_text}\" in {item_split[0]} \"{item_split[1]}\" form"
                                            dirtydata=f"\"{item_question}\" selects \"{codelist_text}\" in {item_split[0]} \"{item_split[1]}\" form" 
                
                
                          
                                
                            details_clean=details_clean.replace(i,cleandata)
                            details_dirty=details_dirty.replace(i,dirtydata)
                            #i_replace的作用是为了统一通过获取双引号内的内容，获取codelist_value，但是在替换时仍然采用i，因为details里的就是i
                  
                            
                        ##logic填充
                        logic=details_clean
                        logic_list=details_clean.split("||")
                        for i in logic_list:
                              i_replace=i.replace("&&",", ",(i.count("&&")-1)).replace("&&"," and ")
                              logic=logic.replace(i,i_replace)    
                        logic=logic.replace("||",", or ")
                        #得到details_clean后，将&&替换成and，||替换成or就是logic
                        #此步作用是，对于多个&&的details，如A&&B&&C||D&&E||F,首先通过||进行拆分得到logic_list=[A&&B&&C,D&&E,F]
                        #1.对于A&&B&&C2个及以上&&的，应该转化为A, BandC(即除最后一个所有&&转化为， ，最后一个&&转化为and)
                        #2.对于D&&E1个&&的，替换成and即可
                        #3.对于F0个and的，不替换
                        #i_replace=i.replace("&&",", ",(i.count("&&")-1)).replace("&&","and")适用以上三种情况
                        if x[f_key[8]]=="Dependency Rule" :
                            triggered_item=f"\"{x[f_key[4]]}\""
                        elif x[f_key[8]]=="show unique details" : 
                            triggered_item=f"\"{x[f_key[2]]} (details)\""
                        elif x[f_key[8]]=="show form" : 
                            triggered_item=f"\"{x[f_key[2]]}\" form"
                        elif x[f_key[8]]=="show visit" :             
                            triggered_item=f"\"{x[f_key[1]]}\" visit"  
                            
                        if x[f_key[-5]]=="Optional":           
                            f.at[x.name,f_key[6]]=f"Only when {logic}, then \"{x[f_key[-6]]}\" is enable."
                        else:
                          f.at[x.name,f_key[6]]=f"Only when {logic}, then {triggered_item} is enable."
                        #如果被触发字段是optional，应该是此字段的cpation can be enable
                        #此步还需进行优化，比如A selects 1 or A selects 2,logic应该为A selects 1 or 2
                        
                        ##cleandata填充
                        f.at[x.name,f_key[15]]=details_clean.replace("||","\n\n").replace("&&","\n")
                        #将details_clean中的||替换成成2个换行符，&&替换成1个换行符，即为clean test data
                
                 
                        ##dirtydata填充
                        if re.search("&&",x[f_key[-3]])==None and re.search("\|\|",x[f_key[-3]])==None:
                            f.at[x.name,f_key[20]]=details_dirty   
                
                        elif re.search("&&",x[f_key[-3]])==None and re.search("\|\|",x[f_key[-3]])!=None:
                            f.at[x.name,f_key[20]]=details_dirty.replace("||","\n")
                            #如果是单一||逻辑符相连，如a||b||c（假设clean为1，dirty为0），dirtydata为a、b、c同时为0
                        elif re.search("&&",x[f_key[-3]])!=None and re.search("\|\|",x[f_key[-3]])==None:
                            f.at[x.name,f_key[20]]=details_dirty.replace("&&","\n\n")
                
                            #如果是单一&&逻辑符相连，如a&&b&&c（假设clean为1，dirty为0），a b c只要其中一个为0就是dirtyvalue
                
                            
                        
                        return x
            
                except:
                    pass
                   # with open(Debug_path, 'a') as debug:
                   #     print(x[-2],file=debug) 
                   
            f.apply(func_dependency,axis=1)
            
            with open(Debug_path, 'a') as debug:
                print("----------------------dependency EC 填充完成----------------------------", file=debug)
            #--------------------------------------------------------------------------------------------------------------------
            if projectName:
                export_path=os.path.join(out_path, projectName + "_" + "TestingCase_V1.0_draft.xlsx")
            else:
                export_path=os.path.join(out_path, "TestingCase_V1.0_draft.xlsx")
            export_path = "C:/Users/YueXu/Desktop/工具代码/testing case 自动化/新建文件夹/TC.xlsx"
            f.to_excel(export_path,index=False,sheet_name="V1.0")
            wb=openpyxl.load_workbook(export_path)
            ws=wb["V1.0"]
            
            ws.column_dimensions["A"].width=5.15
            ws.column_dimensions["B"].width=12.55
            ws.column_dimensions["C"].width=20.84
            ws.column_dimensions["D"].width=11.97
            ws.column_dimensions["E"].width=26.84
            ws.column_dimensions["F"].width=19.82
            ws.column_dimensions["G"].width=40
            ws.column_dimensions["H"].width=30
            ws.column_dimensions["I"].width=10
            ws.column_dimensions["J"].width=8.4  
            ws.column_dimensions["K"].width=9.79  
            ws.column_dimensions["L"].width=11.97
            ws.column_dimensions["M"].width=11.97  
            ws.column_dimensions["N"].width=14.70 
            ws.column_dimensions["O"].width=12.67 
            ws.column_dimensions["P"].width=33
            ws.column_dimensions["Q"].width=13
            ws.column_dimensions["R"].width=2.3
            ws.column_dimensions["S"].width=14.70
            ws.column_dimensions["T"].width=12.67
            ws.column_dimensions["U"].width=33 
            ws.column_dimensions["V"].width=13
            ws.column_dimensions["W"].width=2.3 
            ws.column_dimensions["X"].width=10 
            ws.column_dimensions["Y"].width=10 
            ws.column_dimensions["Z"].width=10 
            ws.column_dimensions["AA"].width=10  
            ws.column_dimensions["AB"].width=5 
            ws.column_dimensions["AC"].width=5 
            ws.column_dimensions["AD"].width=5 
            ws.row_dimensions[1].height=46
            
            for i in ws.iter_rows(min_row=1,min_col=1,max_row=len(f)+1,max_col=29):
                side=Side(style="thin",color="000000")
                for j in i:       
                    j.font=Font(name="Times New Roman",size=11)
                    j.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
                    j.border=Border(left=side,right=side,top=side,bottom=side)
                    
            for i in ws["A1":"AC1"]:
                for j in i:
                    j.font=Font(name="Times New Roman",size=11,bold=True)
                    j.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
                    j.fill=PatternFill(fill_type="solid",fgColor="C0C0C0")
                    
            ws["D1"].fill=PatternFill(fill_type="solid",fgColor="CC6600")
            ws["F1"].fill=PatternFill(fill_type="solid",fgColor="FFFF00")
            ws["J1"].fill=PatternFill(fill_type="solid",fgColor="FFFF00")
            ws["M1"].fill=PatternFill(fill_type="solid",fgColor="FFFF00")
            ws["N1"].fill=PatternFill(fill_type="solid",fgColor="3399FF") 
            ws["O1"].fill=PatternFill(fill_type="solid",fgColor="3399FF") 
            ws["P1"].fill=PatternFill(fill_type="solid",fgColor="3399FF") 
            ws["Q1"].fill=PatternFill(fill_type="solid",fgColor="3399FF")
            ws["R1"].fill=PatternFill(fill_type="solid",fgColor="3399FF")
            ws["S1"].fill=PatternFill(fill_type="solid",fgColor="66CC00")
            ws["T1"].fill=PatternFill(fill_type="solid",fgColor="66CC00")
            ws["U1"].fill=PatternFill(fill_type="solid",fgColor="66CC00")
            ws["V1"].fill=PatternFill(fill_type="solid",fgColor="66CC00")
            ws["W1"].fill=PatternFill(fill_type="solid",fgColor="66CC00")
            ws["X1"].fill=PatternFill(fill_type="solid",fgColor="FF3300")
            
            ws["O1"].value="Test Subject(All visits have been tested)"
            ws["P1"].value="Test Value"
            ws["Q1"].value="Query Fired(Yes, No, Enable, Disable, user-defined)"
            ws["R1"].value="P/F"
            wb.save(export_path)
            pass
        except Exception as e:
            raise Exception(f"Error Occurs: {e}")
        finally:
            if acrf is not None:
                acrf.close()

        