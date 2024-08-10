import os  
import pandas as pd 
import numpy as np 
from openpyxl import Workbook, load_workbook  
from sklearn.metrics import r2_score, mean_absolute_percentage_error  

def kling_gupta_efficiency(y_true, y_pred):
    correlation = np.corrcoef(y_true.flatten(), y_pred.flatten())[0, 1]
    beta = np.std(y_pred) / np.std(y_true)
    alpha = np.mean(y_pred) / np.mean(y_true)
    kge = 1 - np.sqrt((correlation - 1)**2 + (beta - 1)**2 + (alpha - 1)**2)
    return kge

def nash_sutcliffe_efficiency(y_true, y_pred):
    return 1 - (np.sum((y_true - y_pred)**2) / np.sum((y_true - np.mean(y_true))**2))

def wave_hedges_distance(y_true, y_pred):
    return np.sum(np.minimum(y_true, y_pred)) / np.sum(np.maximum(y_true, y_pred))

def vicis_symmetric_distance(y_true, y_pred):
    return np.sum(np.abs(y_true - y_pred)) / np.sum(np.maximum(y_true, y_pred))

def willmotts_agreement_index(y_true, y_pred):
    mean_obs = np.mean(y_true)
    return 1 - (np.sum((y_true - y_pred)**2) / np.sum((np.abs(y_pred - mean_obs) + np.abs(y_true - mean_obs))**2))

def uncertainty_coefficient(y_true, y_pred, level=0.95, resamples=10000):
    from sklearn.utils import resample
    n = len(y_true)
    resampled_correlations = []
    for _ in range(resamples):
        indices = np.random.choice(range(n), n, replace=True)
        resampled_correlations.append(r2_score(y_true[indices], y_pred[indices]))
    lower_bound = np.percentile(resampled_correlations, (1 - level) / 2 * 100)
    upper_bound = np.percentile(resampled_correlations, (1 + level) / 2 * 100)
    return np.mean(resampled_correlations), lower_bound, upper_bound

def root_mean_square_error(y_true, y_pred):
     # Calculate RMSE  
    Errors = y_true - y_pred  
    MSE = np.mean(Errors ** 2)  
    RMSE = np.sqrt(MSE)  
    return RMSE

def compute_all_metrics(y_true, y_pred):  
    return {  
        'R2': r2_score(y_true, y_pred),  
        'RMSE':  root_mean_square_error(y_true, y_pred),
        'MAPE': mean_absolute_percentage_error(y_true, y_pred),  
        'KGE': kling_gupta_efficiency(y_true, y_pred),  
        'NSE': nash_sutcliffe_efficiency(y_true, y_pred),  
        'WHD': wave_hedges_distance(y_true, y_pred),  
        'VSD': vicis_symmetric_distance(y_true, y_pred),  
        'WAI': willmotts_agreement_index(y_true, y_pred),  
    }  

def save_initial_metrics(filename, training_metrics, testing_metrics):  
    wb = None  
    try:  
        if os.path.exists(filename):  
            wb = load_workbook(filename)  
        else:  
            wb = Workbook()  
            wb.active.title = "Metrics"  
            ws = wb.active  

        if 'Metrics' not in wb.sheetnames:  
            ws = wb.create_sheet(title="Metrics")  
        else:  
            ws = wb['Metrics']  

        if ws.max_row == 1 and ws.max_column == 1 and ws['A1'].value is None:  
            ws.append(['Metric', 'Training', 'Testing'])  
        
        for key in training_metrics:  
            ws.append([key, training_metrics[key], testing_metrics[key]])  

        wb.save(filename)  

    except Exception as e:  
        if wb:  
            wb.save(filename)  
        print("Failed to save metrics:", e)  

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, **to_excel_kwargs):  
    try:  
        if os.path.exists(filename):  
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:  
                book = writer.book  
                if sheet_name in book.sheetnames:  
                    startrow = book[sheet_name].max_row if startrow is None else startrow  
                else:  
                    startrow = 0  
                df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, **to_excel_kwargs)  
        else:  
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:  
                df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, **to_excel_kwargs)  
    except Exception as e:  
        print(f"Failed to append data to sheet {sheet_name}:", e)  

def Save_Metrics(y_train, ytr, y_test, yts,metrics_filename):  
    # Compute metrics  
    training_metrics = compute_all_metrics(y_train, ytr)  
    testing_metrics = compute_all_metrics(y_test, yts)  

    train_df = pd.DataFrame({  
        'y_train': y_train.flatten(),  
        'ytr': ytr.flatten()  
    })  

    test_df = pd.DataFrame({  
        'y_test': y_test.flatten(),  
        'yts': yts.flatten()  
    })  


    save_initial_metrics(metrics_filename, training_metrics, testing_metrics)  
    append_df_to_excel(metrics_filename, train_df, sheet_name="Train_Predictions", index=False)  
    append_df_to_excel(metrics_filename, test_df, sheet_name="Test_Predictions", index=False)  

    print("Data Saved Successfully!")  

# Assuming this is within your existing code context  
# y_train, ytr, y_test, yts would be defined during your model training/testing process  

# main(y_train, ytr, y_test, yts)  # Call this function with actual inputs