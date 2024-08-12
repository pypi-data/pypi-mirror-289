import os  
import pandas as pd 
import numpy as np 
from openpyxl import Workbook, load_workbook  
from sklearn.metrics import r2_score, mean_absolute_percentage_error  
from scipy import stats  


def calculate_correlation_coefficient(y_true, y_pred):  
    """  
    Calculate the Pearson correlation coefficient between true and predicted datasets.  

    Parameters:  
    y_true (numpy.ndarray): Observed (true) values.  
    y_pred (numpy.ndarray): Predicted values.  

    Returns:  
    float: Correlation coefficient, with -1 indicating a perfect negative correlation,  
           0 indicating no correlation, and 1 indicating a perfect positive correlation.  
    """  
    # Ensure inputs are numpy arrays  
    y_true = np.asarray(y_true)  
    y_pred = np.asarray(y_pred)  

    # Calculate the means of the arrays  
    mean_y_true = np.mean(y_true)  
    mean_y_pred = np.mean(y_pred)  
    
    # Calculate the numerator and the denominators for the correlation coefficient  
    numerator = np.sum((y_true - mean_y_true) * (y_pred - mean_y_pred))  
    denominator = (np.sum((y_true - mean_y_true) ** 2) ** 0.5) * (np.sum((y_pred - mean_y_pred) ** 2) ** 0.5)  
    
    # Calculate the correlation coefficient  
    R = numerator / denominator  
    
    return R 


def kling_gupta_efficiency(y_true, y_pred):  
    # Ensure the inputs are NumPy arrays to use flatten and other numpy operations  
    if isinstance(y_true, pd.Series):  
        y_true = y_true.values  
    if isinstance(y_pred, pd.Series):  
        y_pred = y_pred.values  

    # Flatten arrays if necessary  
    y_true = y_true.flatten()  
    y_pred = y_pred.flatten()  
    
    # Calculate components of KGE  
    mean_true = np.mean(y_true)  
    mean_pred = np.mean(y_pred)  
    std_true = np.std(y_true)  
    std_pred = np.std(y_pred)  
    
    # Compute correlation coefficient  
    correlation = np.corrcoef(y_true, y_pred)[0, 1]  
    
    # Compute alpha (ratio of standard deviations)  
    alpha = std_pred / std_true  
    
    # Compute beta (ratio of means)  
    beta = mean_pred / mean_true  
    
    # Calculate KGE  
    kge = 1 - np.sqrt((correlation - 1) ** 2 + (alpha - 1) ** 2 + (beta - 1) ** 2)  
    
    return kge 

def nash_sutcliffe_efficiency(y_true, y_pred):
    return 1 - (np.sum((y_true - y_pred)**2) / np.sum((y_true - np.mean(y_true))**2))

def wave_hedges_distance(y_true, y_pred):
    return np.sum(np.minimum(y_true, y_pred)) / np.sum(np.maximum(y_true, y_pred))

def vicis_symmetric_distance(y_true, y_pred):
    return np.sum(np.abs(y_true - y_pred)) / np.sum(np.maximum(y_true, y_pred))

def willmotts_agreement_index(y_true, y_pred):
    mean_obs = np.mean(y_true)
    return 1 - (np.sum((y_true - y_pred)**2) / np.sum((np.abs(y_pred - mean_obs) + np.abs(y_true - mean_obs))**2))

def calculate_u95(y_true, y_pred):  
    """  
    Calculate the 95% confidence interval for the mean of residuals (U95).  

    Parameters:  
    y_true (numpy.ndarray): True observed values.  
    y_pred (numpy.ndarray): Predicted values.  

    Returns:  
    tuple: Lower and upper bounds of the 95% confidence interval.  
    """  
    # Calculate residuals (errors)  
    residuals = y_true - y_pred  
    
    # Calculate the sample mean of residuals  
    mean_residual = np.mean(residuals)  
    
    # Calculate the sample standard deviation of residuals  
    std_residual = np.std(residuals, ddof=1)  # Bessel's correction  
    
    # Sample size  
    n = len(residuals)  
    
    # t-score for 95% confidence level  
    t_score = stats.t.ppf((1 + 0.95) / 2, df=n-1)  
    
    # Standard error of the mean  
    std_error = std_residual / np.sqrt(n)  
    
    # Margin of error  
    margin_of_error = t_score * std_error  
    
    # Confidence interval  
    lower_bound = mean_residual - margin_of_error  
    upper_bound = mean_residual + margin_of_error  
    
    return lower_bound, upper_bound



def root_mean_square_error(y_true, y_pred):
     # Calculate RMSE  
    Errors = y_true - y_pred  
    MSE = np.mean(Errors ** 2)  
    RMSE = np.sqrt(MSE)  
    return RMSE

def compute_all_metrics(y_true, y_pred):
    lower_bound, upper_bound = calculate_u95(y_true, y_pred)
    return {  
        'R': calculate_correlation_coefficient(y_true, y_pred),  
        'RMSE':  root_mean_square_error(y_true, y_pred),
        'MAPE': mean_absolute_percentage_error(y_true, y_pred),  
        'KGE': kling_gupta_efficiency(y_true, y_pred),  
        'NSE': nash_sutcliffe_efficiency(y_true, y_pred),  
        'WHD': wave_hedges_distance(y_true, y_pred),  
        'VSD': vicis_symmetric_distance(y_true, y_pred),  
        'WAI': willmotts_agreement_index(y_true, y_pred),  
        'U95%': f'[{lower_bound:.4f}, {upper_bound:.4f}]'
    }  


def save_initial_metrics(filename, training_metrics, testing_metrics):  
    wb = None  
    try:  
        if os.path.exists(filename):  
            wb = load_workbook(filename)  
        else:  
            wb = Workbook()  
            wb.active.title = "Metrics"  
            ws = wb.active  

        if 'Metrics' not in wb.sheetnames:  
            ws = wb.create_sheet(title="Metrics")  
        else:  
            ws = wb['Metrics']  

        if ws.max_row == 1 and ws.max_column == 1 and ws['A1'].value is None:  
            ws.append(['Metric', 'Training', 'Testing'])  
        
        for key in training_metrics:  
            ws.append([key, training_metrics[key], testing_metrics[key]])  

        wb.save(filename)  

    except Exception as e:  
        if wb:  
            wb.save(filename)  
        print("Failed to save metrics:", e)  

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, **to_excel_kwargs):  
    try:  
        if os.path.exists(filename):  
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:  
                book = writer.book  
                if sheet_name in book.sheetnames:  
                    startrow = book[sheet_name].max_row if startrow is None else startrow  
                else:  
                    startrow = 0  
                df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, **to_excel_kwargs)  
        else:  
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:  
                df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, **to_excel_kwargs)  
    except Exception as e:  
        print(f"Failed to append data to sheet {sheet_name}:", e)  

def Save_Metrics(y_train, ytr, y_test, yts, metrics_filename):  
    # Compute metrics  
    training_metrics = compute_all_metrics(y_train, ytr)  
    testing_metrics = compute_all_metrics(y_test, yts)  

    # Ensure y_train and ytr are NumPy arrays before flattening, if needed  
    y_train_values = y_train.values if isinstance(y_train, pd.Series) else y_train  
    ytr_values = ytr.values if isinstance(ytr, pd.Series) else ytr  

    # Ensure y_test and yts are NumPy arrays before flattening, if needed  
    y_test_values = y_test.values if isinstance(y_test, pd.Series) else y_test  
    yts_values = yts.values if isinstance(yts, pd.Series) else yts  

    train_df = pd.DataFrame({  
        'y_train': y_train_values.flatten(),  
        'ytr': ytr_values.flatten()  
    })  

    test_df = pd.DataFrame({  
        'y_test': y_test_values.flatten(),  
        'yts': yts_values.flatten()  
    })  

    save_initial_metrics(metrics_filename, training_metrics, testing_metrics)  
    append_df_to_excel(metrics_filename, train_df, sheet_name="Train_Predictions", index=False)  
    append_df_to_excel(metrics_filename, test_df, sheet_name="Test_Predictions", index=False)  

    print("Data Saved Successfully!") 
      
      
# Assuming this is within your existing code context  
# y_train, ytr, y_test, yts would be defined during your model training/testing process  

# main(y_train, ytr, y_test, yts)  # Call this function with actual inputs