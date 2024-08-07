import pandas as pd
from openpyxl.utils import get_column_letter
import openpyxl


class ExcelTool:
    def __init__(self) -> None:
        self.width_config = {
            "Comments": {"width": 90},
            "BestSentence1": {"width": 20},
            "BestSentence2": {"width": 20},
            "FeedBack": {"width": 40},
            "timestamp": {"width": 20},
            "level": {"width": 10},
            "topic": {"width": 20},
            "message": {"width": 40},
            "description": {"width": 60},
            "traceback": {"width": 80},
        }

    def setup_width_config(self, **kwargs):
        self.width_config.update(kwargs)
        print(f"Config updated to {self.width_config}")

    def post_excel(self, df: pd.DataFrame, save_path: str, default_width=6):
        if save_path.endswith(".xlsx"):
            save_path = save_path.rstrip(".xlsx")
        writer = pd.ExcelWriter(f"{save_path}.xlsx", engine="openpyxl")
        df.to_excel(writer, index=False, sheet_name="Sheet1")

        # 엑셀 시트 가져오기
        worksheet = writer.sheets["Sheet1"]

        # 열 너비 자동 조정 또는 사용자 지정
        for column in df.columns:
            # column_width = max(df[column].astype(str).apply(len).max(), len(column))
            column_width = default_width
            if column in self.width_config:
                if "width" in self.width_config[column]:
                    column_width = self.width_config[column]["width"]
            worksheet.column_dimensions[
                get_column_letter(df.columns.get_loc(column) + 1)
            ].width = column_width

        # 첫 번째 행 고정
        worksheet.freeze_panes = "A2"

        # 자동 줄 바꿈 적용
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=1,
            max_col=worksheet.max_column,
        ):
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

        # 변경 사항 저장하고 파일 닫기
        writer.close()


import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment


class ExcelHandler:
    def __init__(self, dataframe: pd.DataFrame):
        self.df = dataframe
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        # 데이터프레임을 엑셀 시트에 추가
        for r in dataframe_to_rows(self.df, index=False, header=True):
            self.worksheet.append(r)

    def set_column_width(self, **kwargs):
        """컬럼 너비 설정"""
        for column, width in kwargs.items():
            col_idx = self.df.columns.get_loc(column) + 1
            col_letter = chr(64 + col_idx)
            self.worksheet.column_dimensions[col_letter].width = width

    def freeze_first_row(self):
        """첫 번째 행 고정"""
        self.worksheet.freeze_panes = self.worksheet["A2"]

    def enable_autowrap(self):
        """자동 줄 바꿈 활성화"""
        for row in self.worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

    def add_hyperlink(self, cell, url, display=None):
        """셀에 하이퍼링크 추가"""
        self.worksheet[cell].hyperlink = url
        self.worksheet[cell].value = display if display else url
        self.worksheet[cell].style = "Hyperlink"

    def add_hyperlinks_to_column(self, column_name, urls, display_texts=None):
        """컬럼에 있는 각 셀에 하이퍼링크 추가"""
        if len(urls) != len(self.df):
            raise ValueError(
                "The length of the URL list must match the length of the dataframe."
            )

        if display_texts and len(display_texts) != len(self.df):
            raise ValueError(
                "The length of the display_texts list must match the length of the dataframe."
            )

        col_idx = self.df.columns.get_loc(column_name) + 1
        for i, url in enumerate(urls, start=2):  # start=2 to account for header row
            cell = f"{chr(64 + col_idx)}{i}"
            display = display_texts[i - 2] if display_texts else None
            self.add_hyperlink(cell, url, display)

    def save(self, filename):
        """엑셀 파일 저장"""
        self.workbook.save(filename)

    def close(self):
        """엑셀 파일 닫기"""
        del self.workbook

    def __del__(self):
        self.close()


class DataFrameHandler:
    def __init__(self, dataframe: pd.DataFrame):
        self.df = dataframe

    def filter_rows(self, **conditions):
        """조건에 맞는 행 필터링"""
        query = " & ".join(
            [
                (
                    f"({col}=={repr(val)})"
                    if not isinstance(val, list)
                    else f"({col}.isin({val}))"
                )
                for col, val in conditions.items()
            ]
        )
        return self.df.query(query)

    def group_and_aggregate(self, group_by, **aggregations):
        """그룹화 및 집계"""
        return self.df.groupby(group_by).agg(aggregations)

    def fill_missing(self, strategy="mean", columns=None):
        """결측값 채우기"""
        if columns is None:
            columns = self.df.select_dtypes(include="number").columns
        else:
            columns = self.df[columns].select_dtypes(include="number").columns

        if strategy == "mean":
            self.df[columns] = self.df[columns].fillna(self.df[columns].mean())
        elif strategy == "median":
            self.df[columns] = self.df[columns].fillna(self.df[columns].median())
        elif strategy == "mode":
            self.df[columns] = self.df[columns].fillna(self.df[columns].mode().iloc[0])
        elif strategy == "ffill":
            self.df[columns] = self.df[columns].fillna(method="ffill")
        elif strategy == "bfill":
            self.df[columns] = self.df[columns].fillna(method="bfill")
        return self.df

    def normalize(self, columns=None):
        """정규화"""
        if columns is None:
            columns = self.df.select_dtypes(include="number").columns
        else:
            columns = self.df[columns].select_dtypes(include="number").columns

        self.df[columns] = (self.df[columns] - self.df[columns].mean()) / self.df[
            columns
        ].std()
        return self.df

    def get_dataframe(self):
        """데이터프레임 반환"""
        return self.df


class MyExcel:
    def __init__(self, dataframe: pd.DataFrame):
        self.df_handler = DataFrameHandler(dataframe)
        self.excel_handler = ExcelHandler(self.df_handler.get_dataframe())

    def save(self, filename):
        """엑셀 파일 저장"""
        self.excel_handler.save(filename)

    def close(self):
        """엑셀 파일 닫기"""
        self.excel_handler.close()

    @property
    def df(self):
        """데이터프레임 접근"""
        return self.df_handler.df

    def __getattr__(self, name):
        """속성 접근"""
        if hasattr(self.df, name):
            return getattr(self.df, name)
        raise AttributeError(
            f"'{self.__class__.__name__}' object has no attribute '{name}'"
        )

    # DataFrameHandler 메서드들에 대한 래퍼
    def filter_rows(self, **conditions):
        """행 필터링"""
        return self.df_handler.filter_rows(**conditions)

    def group_and_aggregate(self, group_by, **aggregations):
        """그룹화 및 집계"""
        return self.df_handler.group_and_aggregate(group_by, **aggregations)

    def fill_missing(self, strategy="mean", columns=None):
        """결측값 채우기"""
        return self.df_handler.fill_missing(strategy, columns)

    def normalize(self, columns=None):
        """정규화"""
        return self.df_handler.normalize(columns)

    # ExcelHandler 메서드들에 대한 래퍼
    def set_column_width(self, **kwargs):
        """컬럼 너비 설정"""
        self.excel_handler.set_column_width(**kwargs)

    def freeze_first_row(self):
        """첫 번째 행 고정"""
        self.excel_handler.freeze_first_row()

    def enable_autowrap(self):
        """자동 줄 바꿈 활성화"""
        self.excel_handler.enable_autowrap()

    def add_hyperlink(self, cell, url, display=None):
        """셀에 하이퍼링크 추가"""
        self.excel_handler.add_hyperlink(cell, url, display)

    def add_hyperlinks_to_column(self, column_name, urls, display_texts=None):
        """컬럼에 있는 각 셀에 하이퍼링크 추가"""
        self.excel_handler.add_hyperlinks_to_column(column_name, urls, display_texts)
