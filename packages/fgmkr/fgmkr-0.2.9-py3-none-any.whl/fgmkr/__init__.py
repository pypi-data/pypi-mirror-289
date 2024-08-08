from matplotlib import pyplot as plt
import numpy as np
from scipy import stats
import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from zipfile import BadZipFile

# Saving Data to Excel Sheets
def toSheet(workbook_name, sheetname, **kwargs):
    data_dict = {}
    max_length = 0

    # Generate Dictionary
    for variable_name, values in kwargs.items():
        if not isinstance(values, list):  # list or var
            values = [values]
        if len(values) > max_length:  # Store max length array
            max_length = len(values)
        data_dict[variable_name] = values  # Populate

    # Equal length normalization excel nonsense
    for key in data_dict:
        if len(data_dict[key]) < max_length:
            data_dict[key].extend([None] * (max_length - len(data_dict[key])))

    # Logistics
    df = pd.DataFrame(data_dict)
    current_dir = os.getcwd()
    workbook_path = os.path.join(current_dir, workbook_name)

    # Workbook Sanity Checks
    if os.path.exists(workbook_path):
        try:
            book = load_workbook(workbook_path)
        except (FileNotFoundError, KeyError, BadZipFile):
            book = Workbook()
            book.remove(book.active)  # Remove default sheet
    else:
        book = Workbook()
        book.remove(book.active)  # Remove default sheet

    if sheetname not in book.sheetnames:
        book.create_sheet(title=sheetname)

    # Save the Wb
    book.save(workbook_path)

    # Process
    with pd.ExcelWriter(workbook_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        df.to_excel(writer, sheet_name=sheetname, index=False)

# Single Input
def fgmk(n,x,y,xlabel,ylabel, titlestr, figtext = None, glabel = None, grid = None,dim=None):
    # Figure Setup
    if dim is not None:
        plt.figure(n, figsize=dim)
    else:
        plt.figure(n)

    # Data Visualization
    plt.plot(x,y, label = glabel)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.title(titlestr)
    plt.text(0.5, -0.18, r'$\bf{Figure\ }$' + str(n) + r': '+ str(figtext), transform=plt.gca().transAxes,
            horizontalalignment='center', verticalalignment='center', fontsize=10)

    # Figure Cleanup
    if grid:
        plt.grid(which='both')
    if glabel is not None:
        plt.legend()

# Double Input
def fgmk2(n, x0, y0, xlabel, ylabel, t,
          l=None, lst='-', c='tab:blue',
          x=None, y=None, l1=None, lst1='-', c1='tab:orange',
          fig=None, grid=1, dim=None):
    # Figure Setup
    if dim is not None:
        plt.figure(n, figsize=dim)
    else:
        plt.figure(n)
    plt.title(t)
    if fig is not None:
        plt.text(0.5, -0.18, r'$\bf{Figure\ }$' + str(n) + r': ' + str(fig), transform=plt.gca().transAxes,
                 horizontalalignment='center', verticalalignment='center', fontsize=10)

    # Data Visualization
    plt.plot(x0, y0, label=l, linestyle=lst, color=c)  # Set 0
    if x is not None and y is not None:  # Set 1
        plt.plot(x, y, label=l1, linestyle=lst1, color=c1)
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)

    # Figure Cleanup
    if grid:
        plt.grid(which='both')
    if l is not None or l1 is not None:
        plt.legend()

# Histogram
def hgmk2(n, data, xlabel, ylabel, t, l=None, lst='-', c='tab:blue', fig=None, grid=1, dim=None,
          conf_intervals=None, mark_values=None, mark_labels=None, average=False, bell_curve=False,
          xmax=None):
    # Filter data based on xmax
    if xmax is not None:
        data = data[data <= xmax]

    # Figure Setup
    if dim is not None:
        plt.figure(n, figsize=dim)
    else:
        plt.figure(n)

    # Calculate histogram
    counts, bins, _ = plt.hist(data, bins=30, color='#214187', alpha=0.7, label=l, linestyle=lst, density=True)

    # Average
    if average:
        avg = np.mean(data)
        plt.axvline(avg, color='black', linestyle='-', label=r'Data $\bar{x}$' + f'\n{avg:.3f}')

    # Confidence intervals
    if conf_intervals:
        for ci_value in conf_intervals:
            if ci_value <= 1:
                lower_bound, upper_bound = np.percentile(data, [(1-ci_value)*50, (1+ci_value)*50])
                plt.axvline(lower_bound, color='red', linestyle=':', label=f'{ci_value*100}% \nCI: [{lower_bound:.2f}, {upper_bound:.2f}]')
                plt.axvline(upper_bound, color='red', linestyle=':')
            else:
                lower_bound, upper_bound = np.percentile(data, [ci_value[0], ci_value[1]])
                plt.axvline(lower_bound, color='red', linestyle=':', label=f'{ci_value[0]}% to {ci_value[1]}% \nCI: [{lower_bound:.2f}, {upper_bound:.2f}]')
                plt.axvline(upper_bound, color='red', linestyle=':')

    # Mark values
    if mark_values and mark_labels:
        colors = ['tab:pink', 'tab:orange', 'tab:green', 'tab:red', 'tab:purple', 'tab:brown', 'tab:blue', 'tab:gray']
        for value, label, color in zip(mark_values, mark_labels, colors[:len(mark_values)]):
            plt.axvline(value, color=color, linestyle='--', label=label)
            # plt.text(value, max(counts), label, color=color)

    # Bell curve
    if bell_curve:
        x = np.linspace(min(data), max(data), 100)
        mu = np.mean(data)
        sigma = np.std(data)
        pdf = stats.norm.pdf(x, mu, sigma)
        plt.plot(x, pdf, color='orange', linestyle='-')

    # Legend
    legend = plt.legend()
    for text in legend.get_texts():
        text.set_fontsize(8)

    # Plot settings
    plt.title(t)
    plt.xlabel(xlabel, fontsize=12)
    plt.ylabel(ylabel, fontsize=12)
    plt.xticks(fontsize=10)
    plt.yticks(fontsize=10)
    if grid:
        plt.grid(which='both')
    if fig is not None:
        plt.text(0.5, -0.18, r'$\bf{Figure\ }$' + str(n) + r': ' + str(fig), transform=plt.gca().transAxes,
                 horizontalalignment='center', verticalalignment='center', fontsize=10)
    plt.show()

# Help
def fgmk_help():
    print("fgmk:")
    print("  Single Input:")
    print("    Mandatory:")
    print("      Figure Number, x, y, xlabel, ylabel, titlestr")
    print("    Optional:")
    print("      figtext: Figure text")
    print("      glabel: Graph label")
    print("      grid: Enable grid (default: None)")
    print("      dim: 1x2 Tuple for figure dimensions\n")

    print("toSheet:")
    print("    Mandatory:")
    print("      Excel Name, Sheet Name")
    print("    Optional:")
    print("      Any array or numeric value in dictionary form, such as Var=[21] or Var=Var\n")

    print("fgmk2:")
    print("  Double Input:")
    print("    Mandatory:")
    print("      Figure Number, x0, y0, xlabel, ylabel, t (title)")
    print("    Optional:")
    print("      l: Label for data set 0")
    print("      lst: Line style for data set 0 ('-' by default)")
    print("      c: Color for data set 0 ('tab:blue' by default)")
    print("      x: x values for data set 1")
    print("      y: y values for data set 1")
    print("      l1: Label for data set 1")
    print("      lst1: Line style for data set 1 ('-' by default)")
    print("      c1: Color for data set 1 ('tab:orange' by default)")
    print("      fig: Figure text")
    print("      grid: Enable grid (default: 1)")
    print("      dim: 1x2 Tuple for figure dimensions\n")

    print("hgmk2:")
    print("  Histogram:")
    print("    Mandatory:")
    print("      Figure Number, data, xlabel, ylabel, t (title)")
    print("    Optional:")
    print("      l: Label for the histogram")
    print("      lst: Line style for the histogram ('-' by default)")
    print("      c: Color for the histogram ('tab:blue' by default)")
    print("      fig: Figure text")
    print("      grid: Enable grid (default: 1)")
    print("      dim: 1x2 Tuple for figure dimensions")
    print("      conf_intervals: List of confidence intervals to mark")
    print("      mark_values: Values to mark on the histogram")
    print("      mark_labels: Labels for the marked values")
    print("      average: Plot average line (default: False)")
    print("      bell_curve: Plot bell curve (default: False)")
    print("      xmax: Maximum value for the x-axis\n")
