#!/usr/bin/python3
# -*- coding:utf-8 -*-
"""
@author: JHC
@file: util_excel.py
@time: 2023/5/27 23:37
@desc:
"""
import pandas as pd
from typing import List, Dict
import xlsxwriter
import xlrd
from sdk.base.base_temp import Base
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


class ExcelProcess(Base):
    """
    excel 读写
    """

    def __init__(self):
        super(ExcelProcess, self).__init__()

    def read_yield_by_pandas(self,file,header=None, sheet_name=None):
        """

        :param file:
        :param header:
        :param sheet_name:
        :return:
        """
        if sheet_name:
            sheet_list = [sheet_name]
        else:
            excel_file = pd.ExcelFile(file)
            sheet_list = excel_file.sheet_names
        for sheet_name in sheet_list:
            num = 0
            header_num = 1 if header else header
            df = pd.read_excel(file, header=header_num, sheet_name=sheet_name)
            for row in df.itertuples(index=None, name=None):
                num += 1
                if num == 1:
                    if not header:
                        headers = row
                        continue

                if header:
                    headers = header

                yield {
                    "sheet": sheet_name,
                    "headers": list(headers),
                    "num": num,
                    "line": list(row),
                }


    def read_yield_xlsx(self, file, headers=None, spliter="\t", sheets=None):
        """

        """
        workbook = openpyxl.load_workbook(file)
        if not sheets:
            sheets_obj = workbook.active
            sheets = workbook.sheetnames
        if headers:
            min_row = 1
        else:
            min_row = 2
            headers = [i for i in sheets_obj.iter_rows(min_row=1, max_row=1, values_only=True)][0]
        for index, row in enumerate(sheets_obj.iter_rows(min_row=min_row, values_only=True), start=1):
            yield {
                "headers": headers,
                "sheets": sheets,
                "num": index,
                "line": [str(i) for i in row]
            }

        workbook.close()

    def read_yield(self, file: str, headers: list = None,
                   encoding: str = "utf-8", spliter: str = "\t", sheets: list = None, headersline=[0]) -> dict:
        """
        按行读取excel
        :param file:
        :param headers:[[],[]]每个sheet对应一个header
        :param encoding:
        :param spliter:
        :param sheets:
        :return:
        """
        data = xlrd.open_workbook(file)
        if not sheets:
            sheets = data.sheet_names()
            # sheets = [i for i in data.sheet_names() if "对话" in i]
        # headers = ["i" for i in range(len(sheets))]
        for index, sheet in enumerate(sheets):
            table = data.sheet_by_name(sheet)
            nrows = table.nrows
            # 传headers进来从第1行开始算，不传从第2行开始算
            if not headers:
                # header = []
                for id in headersline:
                    header = table.row_values(id)
                    # header.append(_header)
                    start = 1
            else:
                header = headers[index]
                start = 0
            num = 0
            for row in range(start, nrows):
                info = []
                for i in table.row_values(row):
                    if isinstance(i, str):
                        info.append(i)
                    else:
                        if str(i).endswith(".0"):
                            info.append(str(int(i)))
                        else:
                            info.append(str(i))
                num += 1

                yield {
                    "sheet": sheet,
                    "headers": header,
                    "num": num,
                    "line": info,
                }

    @staticmethod
    def write_exists(file: str, data: List, headers: List, sheets: List = ["Sheet1"], style_header=None):
        """
        向已经存在的excel表中写入数据
        :param file:
        :param data:
        :param headers:
        :param sheets:
        :param style_header: style_header = {
            'bg_color': 'red',
            'font_name': '微软雅黑',
            'font_size': 18,
            'bold': True,
            'align': 'center',
            'valign': 'center',
            'text_wrap': True
        }
        :return:
        """
        if len(data) != len(sheets) or len(headers) != len(sheets):
            raise ValueError("data,headers,sheets长度不一致")
        workbook = openpyxl.load_workbook(file)

        if style_header:
            color_map = {
                'black': '000000',
                'white': 'FFFFFF',
                'red': 'FF0000',
                'green': '00FF00',
                'blue': '0000FF',
                'yellow': 'FFFF00',
                'orange': 'FFA500',
                'purple': '800080',
                'pink': 'FFC0CB',
                'brown': 'A52A2A',
                'gray': '808080'
            }

            font = Font(name=style_header['font_name'], size=style_header['font_size'], bold=style_header['bold'])
            fill = PatternFill(start_color=color_map.get(style_header['bg_color'], "ff2d2d"),
                               end_color=color_map.get(style_header['bg_color'], "ff2d2d"),
                               fill_type="solid")
            alignment = Alignment(horizontal=style_header['align'], vertical=style_header['valign'],
                                  wrap_text=style_header['text_wrap'])
        for index, sheet in enumerate(sheets):
            obj_sheet = workbook.create_sheet(sheet)
            out_data = data[index]
            header = headers[index]

            if header:
                out_data.insert(0, header)

            for row_index, row in enumerate(out_data):
                new_row = [str(i) for i in row]
                obj_sheet.append(new_row)

                if row_index == 0 and style_header:
                    if style_header:
                        for cell in obj_sheet[1]:
                            cell.font = font
                            cell.fill = fill
                            cell.alignment = alignment

        workbook.save(file)
        workbook.close()

    @staticmethod
    def write(file: str, data: List, headers: List, sheets: List = ["Sheet1"], style_header: Dict = None):
        """
        支持同时写入多个sheet
        向已经存在的excel表中写入数据时会覆盖原有数据
        :param file:
        :param data:[[[],[],[]]] 三层
        :param headers:[[],[],[]] 二层
        :param sheets:[,,,] 一层
        :param style_header:{
                        'font_name': '微软雅黑',
                        'font_size': 18,
                        'bold': True,
                        'border': 0,
                        'align': 'center',
                        'valign': 'vcenter',
                        'text_wrap': False,
                        "color":"white",
                        "bg_color":"red",
                        }
        :return:
        """
        if len(data) != len(sheets) or len(headers) != len(sheets):
            raise ValueError("data,headers,sheets长度不一致")
        # 全局样式
        options = {
            'strings_to_numbers': False,
            'strings_to_urls': True,
            'constant_memory': False,
            'default_format_properties': {
                'font_name': '微软雅黑',
                'font_size': 11,
                'bold': False,
                'border': 0,
                'align': 'left',
                'valign': 'vcenter',
                'text_wrap': False,
            },
        }
        workbook = xlsxwriter.Workbook(file, options)
        for index, sheet in enumerate(sheets):
            worksheet = workbook.add_worksheet(sheet)
            data[index].insert(0, headers[index])
            for row, lis in enumerate(data[index]):
                for col, val in enumerate(lis):
                    # header 样式
                    if row == 0:
                        cell_format = workbook.add_format(style_header)
                    else:
                        cell_format = None
                    worksheet.write_string(
                        row=row,
                        col=col,
                        string=str(val),
                        cell_format=cell_format,
                    )
        workbook.close()
