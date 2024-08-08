
import numpy as np
import pandas as pd
import plotly.graph_objects as go
from pyvis.network import Network
from scipy.stats import norm

from openpyxl import Workbook
from openpyxl.utils import get_column_letter as letter
from openpyxl.styles import Font, Color, PatternFill, NamedStyle
from openpyxl.formatting.rule import ColorScale, FormatObject, Rule


# Internal method - prepares data. Expects:
# Example ID column (str), target column (int), prediction column (int), probability vector column (vector).
def __prep_data__(predictions: pd.DataFrame, 
                    id_col: str = "ID", 
                    target_col: str = "TARGET",
                    pred_col: str = "PREDICTION",
                    prob_col: str = "PROBABILITY"):
    
    actual_types = predictions.dtypes.to_dict()
    expected_types = {
        id_col: "object",
        target_col: "int32",
        pred_col: "int32",
        prob_col: "object"
    }

    for col_name, col_type in expected_types.items():
        if col_name not in actual_types:
            raise ValueError(f"The given dataframe is missing the column '{col_name}'!")
        elif actual_types[col_name] != col_type:
            raise ValueError(f"The column {col_name} has incorrect type {actual_types[col_name]}! Expected: {col_type}")

    predictions = predictions[[id_col, target_col, pred_col, prob_col]].rename(columns={
        id_col: "ID",
        target_col: "TARGET",
        pred_col: "PREDICTION",
        prob_col: "PROBABILITY"
    })

    def expand_vec(vec):
        return vec["values"]

    prob_df = pd.DataFrame(predictions["PROBABILITY"].map(expand_vec))
    prob_df = pd.DataFrame(prob_df["PROBABILITY"].to_list())
    prob_df.columns = [str(i) for i in range(len(prob_df.columns))]
    predictions = predictions.drop("PROBABILITY", axis=1)
    predictions = pd.concat([predictions, prob_df], axis=1)

    return predictions



# Class which works with one dataset - saves some repetitive operations by keeping the data.
class ClusterEvaluator:

    # Expects predictions and targets to be integer encoded from 0 to # clusters - 1
    def __init__(self,  
                    predictions: pd.DataFrame, 
                    id_col: str = "ID", 
                    target_col: str = "TARGET",
                    pred_col: str = "PREDICTION",
                    prob_col: str = "PROBABILITY"):

        self.predictions: pd.DataFrame = __prep_data__(predictions, id_col, target_col, pred_col, prob_col)
        self.support: np.ndarray = None
        self.num_clusters: int = predictions["TARGET"].nunique()
        self.count = predictions.shape[0]
        self.targets_arr = self.predictions[["TARGET"]].to_numpy().reshape(-1).astype(int)
        self.off_prob_arr = self.predictions.drop(["ID", "TARGET", "PREDICTION"], axis=1).to_numpy()
        self.off_prob_arr[np.arange(self.count), self.targets_arr] = 0



    # Get the influence count matrix. For all classes i, j: 
    # Among examples of target class i, how many of them have probability in j above a certain threshold?
    # result[i][j] = influence count of j on target i.
    def get_influence_counts(self, detection_thresh: float = 0.05) -> np.ndarray:

        agg_dict = {str(i): (lambda x: (x >= detection_thresh).sum()) for i in range(self.num_clusters)}
        influence_counts = self.predictions.groupby("TARGET").agg(agg_dict).reset_index()
        influence_counts = influence_counts.sort_values(by="TARGET")

        return influence_counts.drop("TARGET", axis=1).to_numpy()
    

    # Gets the support of each target, ordered from 0 upwards.
    def get_support(self) -> np.ndarray:

        if self.support is not None:
            return self.support
        
        support_df = self.predictions.groupby("TARGET").size().reset_index(name='count').sort_values(by="TARGET")
        self.support = support_df[["count"]].to_numpy().reshape(-1)

        return self.support
    

    # Influence = influence count normalized by support of class i. Rows do NOT add to 1.
    # Strictly speaking, an influence[i][j] of 5% means that class j affects 5% of the data with target i.
    # Multiple clusters can affect the same data points, which is why rows don't add to 1.
    def get_influence(self, detection_thresh: float = 0.05) -> np.ndarray:

        counts = self.get_influence_counts(detection_thresh)
        support = self.get_support().reshape(-1, 1)

        return counts / support


    # Creates a dictionary from the influence matrix, filtering only the links passing the influence_thresh.
    # print parameter indicates whether to print phrases or not.
    def get_influence_dictionary(self, 
                                 detection_thresh: float = 0.05, 
                                 influence_thresh: float = 0.02,
                                 printPhrases = True) -> dict:
        
        # Set diagonal to 0 so that it isn't counted. (Cluster does not "influence" itself).
        influence_matrix = self.get_influence(detection_thresh)
        np.fill_diagonal(influence_matrix, 0)

        influence_counts = self.get_influence_counts(detection_thresh)
        support_arr = self.get_support()

        influence_dict = {}

        for i in range(self.num_clusters):
            for j in range(self.num_clusters):
                if influence_matrix[i][j] >= influence_thresh:
                    influence_dict[(i, j)] = (influence_matrix[i][j], influence_counts[i][j], support_arr[i])

        # Optionally print readable phrases.
        if printPhrases:
            influence_dict_sorted = sorted(influence_dict.items(), key=lambda item: item[1][0], reverse=True)
            for item in influence_dict_sorted:
                start, end = item[0]
                prob, count, support = item[1]
                print(f"Cluster {end} influences cluster {start} by {prob:.2%}  ({count} / {support})")   

        return influence_dict
    

    # Create a graph out of the edges that appear in the influence dictionary above.
    def create_influence_graph(self,
                               detection_thresh: float  = 0.05,
                               influence_thresh: float = 0.02,
                               filename: str = "clustering_influence_graph.html"):

        influence_matrix = self.get_influence(detection_thresh)
        np.fill_diagonal(influence_matrix, 0) # No self-edges in the graph!

        # Create a directed graph
        net = Network(notebook=True, directed=True, height="750px", width="100%", bgcolor="#222222", font_color="black")

        # Add nodes.
        for i in range(self.num_clusters):
            net.add_node(i, label=str(i), shape="circle")

        # Function to determine color brightness based on weight
        def get_color(weight):
            weight = weight / 0.05
            base_color = "255,255,255"  # White in RGB
            intensity = max(0, min(255, int(255 * weight)))  # Adjust intensity based on weight
            return f"rgba({base_color},{intensity})"

        # Add edges with weights, adjusting for bidirectional edges to avoid overlap
        for i in range(self.num_clusters):
            for j in range(self.num_clusters):
                if influence_matrix[i][j] >= influence_thresh:
                    color = get_color(influence_matrix[i][j])
                    edge_label = f"{influence_matrix[i][j]:.2%}"
                    if influence_matrix[j][i] >= influence_thresh:  # Check for bidirectional edge
                        # For bidirectional edges, use "smooth" type to avoid overlap
                        net.add_edge(i, j, title=edge_label, label=edge_label, width=3, 
                                     smooth={'type': 'curvedCW', 'roundness': 0.2}, arrowStrikethrough=False, color=color)
                    else:
                        # For unidirectional edges, use straight lines
                        net.add_edge(i, j, title=edge_label, label=edge_label, width=3, arrowStrikethrough=False, color=color)

        # Set options for the graph to make it more interactive
        net.set_options(
        """
        var options = {
        "nodes": {
            "shape": "circle",
            "size": 50,
            "font": {
            "size": 18,
            "face": "arial",
            "strokeWidth": 0,
            "align": "center",
            "color": "black"
            }
        },
        "edges": {
            "arrows": {
            "to": {
                "enabled": true,
                "scaleFactor": 1
            }
            },
            "color": {
            "inherit": false
            },
            "smooth": {
            "enabled": true,
            "type": "dynamic"
            },
            "font": {
            "size": 14,
            "align": "top",
            "strokeWidth": 0,
            "color": "white"
            },
            "width": 2
        },
        "physics": {
            "forceAtlas2Based": {
            "gravitationalConstant": -100,
            "centralGravity": 0.01,
            "springLength": 100,
            "springConstant": 0.08
            },
            "maxVelocity": 50,
            "solver": "forceAtlas2Based",
            "timestep": 0.35,
            "stabilization": {
            "enabled": true,
            "iterations": 1000,
            "updateInterval": 25,
            "onlyDynamicEdges": false,
            "fit": true
            }
        }
        }
        """
        )

        net.save_graph(filename)


    # Gets the total leakage: the number of examples with off-probability more than the threshold.
    def get_total_leakage(self, detection_thresh: float = 0.05) -> float:

        max_off_prob = np.max(self.off_prob_arr, axis=1)
        num_leaks = np.sum(max_off_prob >= detection_thresh)

        return num_leaks / self.count
    

    # Do hypothesis testing on the total leakage at different detection thresholds and different comparison values
    # for the alternative hypothesis. Generate a plotly table, also returns data behind the table in a tuple.
    def hypothesis_test_total_leakage(self, significance_level = 0.05):

        comparison_values = np.linspace(0.05, 0.25, num=9)
        detection_thresholds = np.linspace(0.05, 0.25, num=9)

        xlen = len(detection_thresholds)
        ylen = len(comparison_values)

        total_leakage_stats = np.zeros(xlen)
        for i in range(xlen):
            total_leakage_stats[i] = self.get_total_leakage(detection_thresholds[i])

        # The null hypothesis is "the total leakage is equal to X"
        # The alternative is "the total leakage is more than X"
        # Performing the grid of hypothesis tests:

        alpha = significance_level

        n = self.count
        p_0 = comparison_values.reshape(-1, 1)
        p_hat = total_leakage_stats.reshape(1, -1)

        numerator = np.repeat(p_hat, ylen, axis=0) - np.repeat(p_0, xlen, axis=1)
        denominator = np.sqrt(p_0 * (1 - p_0) / n)
        denominator = np.repeat(denominator, xlen, axis=1)
        z_stats = numerator / denominator

        p_values = 1 - norm.cdf(z_stats)
        decisions = p_values < alpha

        # Format decisions in a table.

        cell_labels = np.round(np.repeat(p_hat, ylen, axis=0), decimals=4)
        cell_colors = decisions.astype(int)
        x_axis = np.round(detection_thresholds, decimals=4)
        y_axis = np.round(comparison_values, decimals=4)

        fig = go.Figure(data=go.Heatmap(
            z=cell_colors,
            text=cell_labels,
            texttemplate="%{text}",
            colorscale=[[0, "green"], [1, "red"]],
            showscale=False  # This line removes the colorbar
        ))

        fig.update_layout(
            title="Hypothesis Testing for Cluster Leakage",
            xaxis=dict(title="Detection Threshold", tickvals=list(range(len(x_axis))), ticktext=x_axis),
            yaxis=dict(title="Comparison", tickvals=list(range(len(y_axis))), ticktext=y_axis)
        )

        return (fig, cell_labels, cell_colors, x_axis, y_axis)
    
    
    # Write full report to xlsx. (Does not include influence graph)
    def save_xml_report(self, 
                        detection_thresh: float = 0.05, 
                        influence_thresh: float = 0.02,
                        significance_level: float = 0.05,
                        filename: str = "cluster_leakage_report.xlsx"):

        col_names = ["ID", "TARGET", "PREDICTION"]
        predictions_arr = self.predictions.to_numpy()
        influence_arr = self.get_influence(detection_thresh)
        _, cell_labels, cell_colors, x_axis, y_axis = self.hypothesis_test_total_leakage(significance_level)

        wb = Workbook()

        # ---------------------------------------------------------------------   Write to Sheet 1.
        sheet1 = wb.active
        sheet1.title = "Predictions and Targets"

        # Header row
        sheet1.freeze_panes = "A2" # Freeze header row.
        header = col_names + [f"CLUSTER_{i}" for i in range(self.num_clusters)]
        sheet1.append(header)

        header_style = NamedStyle(name="header")
        header_style.font = Font(bold=True)
        header_style.fill = PatternFill(patternType="solid", fgColor="9ec0ff")

        for col in sheet1.iter_cols(min_row=1, max_row=1):
            col[0].style = header_style

        # Add the prediction data.
        for i in range(predictions_arr.shape[0]):
            sheet1.append(list(predictions_arr[i]))

        last_row = 1 + predictions_arr.shape[0]
        last_col = letter(sheet1.max_column)
        last_cell_address = f"{last_col}{last_row}"

        percent_style = NamedStyle(name="percent")
        percent_style.number_format = "0.00%"
        cells = sheet1[f"D2:{last_cell_address}"]
        for row in cells:
            for cell in row:
                cell.style = percent_style

        # Conditional color gradient formatting for the probability outputs.
        first = FormatObject(type='num', val=0)
        last = FormatObject(type='num', val=1)
        colors = [Color('FFFFFF'), Color('02bd56')]
        color_scale = ColorScale(cfvo=[first, last], color=colors)
        color_rule = Rule(type='colorScale', colorScale=color_scale)
        sheet1.conditional_formatting.add(f"D2:{last_cell_address}", color_rule)
    
        # -----------------------------------------------------------------------------  Write to sheet 2.
        sheet2 = wb.create_sheet("Cluster Leakage")

        # Cluster Title Row
        cluster_names = [f"CLUSTER_{i}" for i in range(self.num_clusters)]
        i = 0
        for col in sheet2.iter_cols(min_row=1, max_row=1, min_col=4, max_col=(4 + self.num_clusters - 1)):
            col[0].value = cluster_names[i]
            col[0].style = header_style
            i = i + 1

        # Statistic Titles
        cells = sheet2["C2":"C6"]
        titles = ["SUPPORT", "COUNT > 5%", "LEAKY RECALL", "RECALL", "LEAKAGE"]
        i = 0
        for row in cells:
            for cell in row:
                cell.value = titles[i]
                cell.style = header_style
                i = i + 1

        # Fill in support.
        i = 0
        for col in sheet2.iter_cols(min_row=2, max_row=2, min_col=4, max_col=(4 + self.num_clusters - 1)):
            col[0].value = f"=COUNTIF('Predictions and Targets'!B:B, \"={i}\")"
            i = i + 1

        # Fill in Count > 5%.
        i = 0
        for col in sheet2.iter_cols(min_row=3, max_row=3, min_col=4, max_col=(4 + self.num_clusters - 1)):
            col[0].value = f"=COUNTIF('Predictions and Targets'!{letter(4+i)}:{letter(4+i)}, \">0.05\")"
            i = i + 1

        # Fill in Leaky Recall.
        i = 0
        for col in sheet2.iter_cols(min_row=4, max_row=4, min_col=4, max_col=(4 + self.num_clusters - 1)):
            col[0].value = f"={letter(4+i)}3 / {letter(4+i)}2"
            col[0].style = percent_style
            i = i + 1

        # Fill in Recall.
        i = 0
        for col in sheet2.iter_cols(min_row=5, max_row=5, min_col=4, max_col=(4 + self.num_clusters - 1)):
            col[0].value = f"=COUNTIFS('Predictions and Targets'!B:B, \"={i}\", 'Predictions and Targets'!C:C, \"={i}\") / {letter(4+i)}2"
            col[0].style = percent_style
            i = i + 1

        # Fill in Leakage percentage.
        i = 0
        for col in sheet2.iter_cols(min_row=6, max_row=6, min_col=4, max_col=(4 + self.num_clusters - 1)):
            col[0].value = f"={letter(4+i)}4 - {letter(4+i)}5"
            col[0].style = percent_style
            i = i + 1

        # Create axes for leakage matrix.
        axes_label_fill = PatternFill(patternType="solid", fgColor="ed9fd4")
        for col in sheet2.iter_cols(min_row=11, max_row=11, min_col=4, max_col=(4 + self.num_clusters - 1)):
            col[0].fill = axes_label_fill
        for row in sheet2.iter_rows(min_row=13, max_row=(13 + self.num_clusters - 1), min_col=2, max_col=2):
            row[0].fill = axes_label_fill
        sheet2["F11"] = "LEAKS TO CLUSTER"
        sheet2["B15"] = "TARGET CLUSTER"
        i = 0
        for col in sheet2.iter_cols(min_row=12, max_row=12, min_col=4, max_col=(4 + self.num_clusters - 1)):
            col[0].value = cluster_names[i]
            col[0].style = header_style
            i = i + 1
        i = 0
        for row in sheet2.iter_rows(min_row=13, max_row=(13 + self.num_clusters - 1), min_col=3, max_col=3):
            row[0].value = cluster_names[i]
            row[0].style = header_style
            i = i + 1

        # Add the inter-cluster leakages.
        matrix_corner = f"{letter(4 + self.num_clusters - 1)}{13 + self.num_clusters - 1}"
        cells = sheet2["D13":matrix_corner]
        for i in range(len(cells)):
            for j in range(len(cells[0])):
                cells[i][j].value = influence_arr[i][j]
                cells[i][j].style = percent_style

        # Delete the diagonal of the matrix.
        black_fill = PatternFill(patternType="solid", fgColor="000000")
        for i in range(self.num_clusters):
            address = f"{letter(4 + i)}{13 + i}"
            sheet2[address] = ""
            sheet2[address].fill = black_fill

        # Gradient conditional formatting
        first = FormatObject(type='num', val=influence_thresh)
        last = FormatObject(type='max')
        colors = [Color('FFFFFF'), Color('02bd56')]
        color_scale = ColorScale(cfvo=[first, last], color=colors)
        color_rule = Rule(type='colorScale', colorScale=color_scale)
        sheet2.conditional_formatting.add(f"D13:{matrix_corner}", color_rule)

        # --------------------------------------------------------------------------------- Write to Sheet 3.
        sheet3 = wb.create_sheet("Total Leakage Tests")

        # Create axes for leakage matrix.
        for col in sheet3.iter_cols(min_row=2, max_row=2, min_col=4, max_col=(4 + 9 - 1)):
            col[0].fill = axes_label_fill
        for row in sheet3.iter_rows(min_row=4, max_row=(4 + 9 - 1), min_col=2, max_col=2):
            row[0].fill = axes_label_fill
        sheet3["F2"] = "DETECTION THRESHOLD"
        sheet3["B6"] = "COMPARISON"
        i = 0
        for col in sheet3.iter_cols(min_row=3, max_row=3, min_col=4, max_col=(4 + 9 - 1)):
            col[0].value = x_axis[i]
            col[0].style = header_style
            i = i + 1
        i = 0
        for row in sheet3.iter_rows(min_row=4, max_row=(4 + 9 - 1), min_col=3, max_col=3):
            row[0].value = y_axis[i]
            row[0].style = header_style
            i = i + 1

        # Fill in the matrix with the total leakage and hypotheses.
        matrix_corner = f"{letter(4 + 9 - 1)}{4 + 9 - 1}"
        cells = sheet3["D4":matrix_corner]
        NO_fill = PatternFill(patternType="solid", fgColor="76e393")
        YES_fill = PatternFill(patternType="solid", fgColor="e37676")
        for i in range(len(cells)):
            for j in range(len(cells[0])):
                cells[i][j].value = cell_labels[i][j]
                cells[i][j].style = percent_style
                cells[i][j].fill = NO_fill if cell_colors[i][j] == 0 else YES_fill

        # Small key for coloring.
        sheet3["D15"] = "REJECT"
        sheet3["D15"].fill = YES_fill
        sheet3["D16"] = "FAIL TO REJECT"
        sheet3["D16"].fill = NO_fill
        sheet3["E15"] = "Statistically significant evidence that cell value (total leakage) is larger than comparison value."
        sheet3["E16"] = "No evidence that cell value (total leakage) is larger than comparison value."

        # -------------------------------------------------------------------- Cleanup and Save.

        # Iterate over all columns and adjust their widths
        for ws_name in wb.sheetnames:
            for column in wb[ws_name].columns:
                column_letter = column[0].column_letter
                wb[ws_name].column_dimensions[column_letter].width = 15

        wb.save(filename)